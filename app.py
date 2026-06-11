"""JPMorgan Chase — Systemically Important Banks Dashboard.

Heroku-ready Dash app. Key architecture notes:
  * NON-BLOCKING BOOT: the FDIC fetch (19 banks, full history) runs in a
    background thread started on the first page request. The web dyno binds
    to $PORT instantly, so Heroku's 60s R10 boot timeout and gunicorn's
    worker timeout can never kill the app while data loads. Visitors see a
    branded loading screen that live-updates fetch progress and auto-reloads
    into the dashboard when ready.
  * REAL DATA ONLY: no synthetic/sample fallback. If the FDIC API cannot
    supply a usable dataset, an error screen is served with a retry link.
  * Run with:  gunicorn app:server --workers 1 --threads 8 --timeout 120
    (one worker: the in-process loader thread + file cache should not be
    duplicated across workers; threads serve concurrent users.)
"""

import base64
import requests
import pandas as pd
import numpy as np
from typing import List, Dict, Union, Optional, Tuple, Any
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import dash
from dash import dcc, html, Output, Input, State, no_update
import dash_bootstrap_components as dbc
from datetime import datetime
from scipy import stats
import logging
import json
import os
import io
import ssl
import glob
import random
import hashlib
import threading
import time
from dash.exceptions import PreventUpdate
from flask import request as flask_request

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
try:
    ssl._create_default_https_context = ssl._create_unverified_context
except AttributeError:
    pass

BASE_URL = "https://banks.data.fdic.gov/api"
DEFAULT_START_DATE = '20030331'
DEFAULT_END_DATE = datetime.today().strftime('%Y%m%d')
REQUESTED_START_DATE_DISPLAY = '03/31/2003'
COMMON_FULL_PEER_START_DATE_DISPLAY = '12/31/2008'
CACHE_DIR = 'data_cache'
os.makedirs(CACHE_DIR, exist_ok=True)
PRIMARY_BANK_DISPLAY_NAME = "JPMorgan Chase"
PRIMARY_BANK_ABBR = "JPM"
PRIMARY_BANK_FDIC_NAME = "JPMorgan Chase Bank, National Association"
DASHBOARD_TITLE = "JPMorgan Chase \u2014 Systemically Important Banks Dashboard"
DASHBOARD_SHORT_TITLE = "SIB Dashboard"
PEER_UNIVERSE_LABEL = "Systemically Important Banks (SIBs)"
HEADER_DISCLOSURE_SHORT = "Personal project \u00b7 public FDIC data \u00b7 not affiliated with JPMorgan Chase"
FOOTER_DISCLOSURE_NOTE = "Independent personal project using public FDIC data; not affiliated with or endorsed by JPMorgan Chase."
PAIRED_GRAPH_HEIGHT = 340
PAIRED_CARD_MIN_HEIGHT = 432
OVERVIEW_GAUGE_SIZE = 78

GRAPH_CONFIG = {
    'displayModeBar': False,
    'scrollZoom': False,
    'doubleClick': False,
    'showTips': False,
    'responsive': True,
}

PRIOR_PERIOD_TOLERANCE_DAYS = 45

# =============================================================================
# CECL TRANSITION ADJUSTMENT TO THE CREDIT-CONCENTRATION DENOMINATOR
# -----------------------------------------------------------------------------
# WHY THIS EXISTS
#   Supervisory credit-concentration ratios (UBPR Page 7B, the "... to Tier 1 +
#   ACL" family) divide an exposure by Tier 1 capital plus the ACL attributable
#   to loans & leases (Fed SR 20-8). For banks that elected the 2019 or 2020
#   CECL capital transition, a portion of the ACL is added back INTO Tier 1
#   (through retained earnings) during the transition. If you then add the FULL
#   ACL on top of an unadjusted Tier 1, you double-count the reserve. To prevent
#   that, the supervisory methodology SUBTRACTS the transitional add-back from
#   Tier 1 in the concentration denominator (OCC Bulletin 2020-90; OCC
#   "Concentrations of Credit" booklet, Oct 2020).
#
#   An unadjusted Tier 1 + ACL denominator therefore runs slightly HIGH, which
#   makes the ratios run slightly LOW. Empirically vs UBPR for JPM (cert 628):
#   denominator ~1.9% high at 03/2020, ~0.4% high at 03/2023, and exactly
#   correct from 2025 onward (transition fully phased out). The effect is a pure
#   denominator scalar, identical across every concentration line for a quarter.
#
# THE ADD-BACK IS NOT ESTIMATED
#   add-back T = (RC-R Part I, line 2 retained earnings)
#                - (RC Schedule, item 26.a retained earnings)
#   Per the FFIEC Call Report instructions, RC-R Part I line 2 retained earnings
#   is defined as RC 26.a retained earnings PLUS the bank's applicable phase-in
#   percentage of its CECL transitional amount. So the difference already embeds
#   whichever election (3-yr 2019: 75/50/25; 5-yr 2020/CARES: 100/100/75/50/25)
#   and whichever year the bank is in -- nothing about the schedule is modeled
#   here. Pre-adoption and 2025+, the difference is ~0 and this is a no-op.
#
# SOURCING (field names confirmed against risview_properties.yaml, the FDIC
# financials field dictionary)
#   The FDIC financials endpoint exposes NO CECL/transition field, but it DOES
#   expose the two retained-earnings stocks the OCC formula needs:
#       RBCEQUP = "RETAINED EARNINGS - RBC"  -> RC-R Part I line 2 (regulatory;
#                  includes the applicable % of the CECL transitional amount).
#       EQUPGR  = "UNDIVIDED PROFITS"        -> Schedule RC item 26.a (book RE).
#   add-back T = RBCEQUP - EQUPGR  (>=0 in transition, ~0 otherwise). Both are in
#   $000s, consistent with RBCT1J (Tier 1) and LNATRES (ACL). They are pulled in
#   the MAIN financials projection (see BankDataRepository.FF) -- no extra calls.
#   Behavior:
#     * Fields present for a row -> exact, UBPR-aligned concentration ratios.
#     * Fields null/absent       -> SAFE NO-OP for that row: raw Tier 1 + ACL
#                                   (current behavior); banner reports coverage.
#   NOTE: risview descriptions are blank, so RBCEQUP/EQUPGR are mapped from their
#   titles. The acceptance test settles it: you have the UBPR Page 7B file for
#   cert 628 -- with these live, the dashboard's 03/2020 and 03/2023 NDFI values
#   should hit 88.21 and 93.66 (every 7B line in proportion); 2025+ already
#   matches. If they don't, the semantics differ -- re-check before trusting it.
#   (Raw FFIEC CDR fallbacks if ever needed: RC 26.a = MDRM 3247; RC-O Memo 5,
#   2021Q2+, = the L&L-attributable add-back MW53.)
# =============================================================================
APPLY_CECL_TRANSITION_ADJUSTMENT = True
CECL_TRANSITION_START = '20200101'   # first CECL adoption period (inclusive)
CECL_TRANSITION_END = '20241231'     # last period with a nonzero add-back (inclusive)

# No single direct add-back field exists in the FDIC financials API, so the
# add-back is derived from the two retained-earnings stocks below.
CECL_DIRECT_ADDBACK_FIELDS: Tuple[str, ...] = ()

# CONFIRMED field names (risview_properties.yaml):
#   RBCEQUP -> RC-R Part I line 2 regulatory retained earnings (transition-incl.)
#   EQUPGR  -> Schedule RC item 26.a book retained earnings (undivided profits)
# First field that resolves on a row is used; absent -> safe no-op for that row.
CECL_REG_RE_FIELDS: Tuple[str, ...] = ("RBCEQUP",)
CECL_BOOK_RE_FIELDS: Tuple[str, ...] = ("EQUPGR",)

# These two fields ride in the MAIN financials projection (BankDataRepository.FF),
# so the separate aux request is unnecessary. The aux machinery is left in place
# but inert (empty field list -> _enrich_cecl returns immediately).
CECL_AUX_FIELDNAMES: Tuple[str, ...] = ()
CECL_AUX_FF = ""

CS = {
    'primary': '#005EB8', 'primary_light': '#2F7FD3', 'primary_dark': '#003B73',
    'secondary': '#333333', 'accent': '#0B4F8A', 'accent_light': '#E8F2FC',
    'bg': '#f4f6f9', 'bg_elevated': '#ffffff',
    'text': '#0f172a', 'text2': '#475569', 'text3': '#64748b',
    'light': '#94a3b8', 'lighter': '#cbd5e1',
    'grid': 'rgba(15,23,42,0.05)', 'border': 'rgba(15,23,42,0.06)',
    'border_strong': 'rgba(15,23,42,0.12)',
    'ghb': '#005EB8', 'peer': '#94a3b8', 'peer_op': 0.55,
    'good': '#16a34a', 'good_light': '#dcfce7', 'good_dark': '#166534',
    'warn': '#f59e0b', 'warn_light': '#fef3c7', 'warn_dark': '#b45309',
    'bad': '#ef4444', 'bad_light': '#fee2e2', 'bad_dark': '#b91c1c',
    'neutral': '#64748b', 'neutral_light': '#f1f5f9',
    'peer_band_top': '#475569', 'peer_band_mid': '#64748b', 'peer_band_low': '#94a3b8',
    'peer_band_bg': '#f8fafc', 'peer_tint': 'rgba(100,116,139,0.06)',
    'hover_bg': '#f8fafc', 'ghb2': '#0B4F8A',
    'gold': '#d4a017', 'silver': '#a8a8a8', 'bronze': '#cd7f32',
    'spark': '#005EB8', 'spark_area': 'rgba(0,94,184,0.08)',
}

# Bumped so caches written before the segment-level asset-quality fields existed
# are invalidated and refetched (the financials projection now also carries the
# RC-N nonaccrual / 90+ / 30-89 detail, segment net charge-offs, and the CAVG5
# average-balance denominators). Also retains the CECL retained-earnings fields.
CACHE_SCHEMA_VERSION = "v13_sib_jpm_segment_assetquality_nco_na_p9_p3_cecl"

BANK_INFO = [
    {"cert": "628",   "display": "JPMorgan Chase"},
    {"cert": "3510",  "display": "Bank of America"},
    {"cert": "3511",  "display": "Wells Fargo"},
    {"cert": "7213",  "display": "Citigroup"},
    {"cert": "33124", "display": "Goldman Sachs"},
    {"cert": "32992", "display": "Morgan Stanley"},
    {"cert": "639",   "display": "BNY Mellon"},
    {"cert": "14",    "display": "State Street"},
    {"cert": "6548",  "display": "U.S. Bancorp"},
    {"cert": "6384",  "display": "PNC"},
    {"cert": "9846",  "display": "Truist"},
    {"cert": "4297",  "display": "Capital One"},
    {"cert": "6672",  "display": "Fifth Third"},
    {"cert": "12368", "display": "Regions Financial"},
    {"cert": "6560",  "display": "Huntington"},
    {"cert": "588",   "display": "M&T Bank"},
    {"cert": "57957", "display": "Citizens Financial"},
    {"cert": "17534", "display": "KeyCorp"},
    {"cert": "57803", "display": "Ally Financial"},
]

CERT_TO_DISPLAY = {b["cert"]: b["display"] for b in BANK_INFO}

# The primary bank's cert is the one hard requirement for the dashboard to
# render; derived from BANK_INFO so it can never drift out of sync.
PRIMARY_BANK_CERT = next(b["cert"] for b in BANK_INFO if b["display"] == PRIMARY_BANK_DISPLAY_NAME)

BANK_NAME_MAPPING = {
    "JPMORGAN CHASE BANK, NATIONAL ASSOCIATION": "JPMorgan Chase",
    "JPMORGAN CHASE BANK, N.A.": "JPMorgan Chase",
    "BANK OF AMERICA, NATIONAL ASSOCIATION": "Bank of America",
    "BANK OF AMERICA, N.A.": "Bank of America",
    "WELLS FARGO BANK, NATIONAL ASSOCIATION": "Wells Fargo",
    "WELLS FARGO BANK, N.A.": "Wells Fargo",
    "CITIBANK, NATIONAL ASSOCIATION": "Citigroup",
    "CITIBANK, N.A.": "Citigroup",
    "GOLDMAN SACHS BANK USA": "Goldman Sachs",
    "MORGAN STANLEY BANK, NATIONAL ASSOCIATION": "Morgan Stanley",
    "MORGAN STANLEY BANK, N.A.": "Morgan Stanley",
    "THE BANK OF NEW YORK MELLON": "BNY Mellon",
    "BANK OF NEW YORK MELLON, THE": "BNY Mellon",
    "STATE STREET BANK AND TRUST COMPANY": "State Street",
    "U.S. BANK NATIONAL ASSOCIATION": "U.S. Bancorp",
    "US BANK NATIONAL ASSOCIATION": "U.S. Bancorp",
    "PNC BANK, NATIONAL ASSOCIATION": "PNC",
    "PNC BANK, N.A.": "PNC",
    "TRUIST BANK": "Truist",
    "CAPITAL ONE, NATIONAL ASSOCIATION": "Capital One",
    "CAPITAL ONE, N.A.": "Capital One",
    "FIFTH THIRD BANK, NATIONAL ASSOCIATION": "Fifth Third",
    "FIFTH THIRD BANK, N.A.": "Fifth Third",
    "REGIONS BANK": "Regions Financial",
    "THE HUNTINGTON NATIONAL BANK": "Huntington",
    "HUNTINGTON NATIONAL BANK, THE": "Huntington",
    "HUNTINGTON NATIONAL BANK": "Huntington",
    "MANUFACTURERS AND TRADERS TRUST COMPANY": "M&T Bank",
    "CITIZENS BANK, NATIONAL ASSOCIATION": "Citizens Financial",
    "CITIZENS BANK, N.A.": "Citizens Financial",
    "KEYBANK NATIONAL ASSOCIATION": "KeyCorp",
    "KEYBANK N.A.": "KeyCorp",
    "ALLY BANK": "Ally Financial",
}

EXECUTIVE_KPIS = [
    ('Return on Assets', 'Profitability'),
    ('Net Interest Margin', 'Margin'),
    ('Efficiency Ratio', 'Efficiency'),
    ('Leverage (Core Capital) Ratio', 'Capital'),
    ('Nonaccrual / Total Loans', 'Credit'),
    ('Net Loan Growth Rate', 'Growth'),
]

# =============================================================================
# SEGMENT-LEVEL ASSET QUALITY
# (Net Charge-Offs / Nonaccrual / 90+ DPD / 30-89 DPD, broken out by loan type)
# -----------------------------------------------------------------------------
# SINGLE SOURCE OF TRUTH. Each tuple supplies the FDIC fields needed to build,
# for one loan segment, the full asset-quality picture that the Piermont
# stability builder produces per tab -- but here for all 19 SIB peers at once:
#
#   * STOCK rates (point-in-time): Nonaccrual, 90+ DPD still accruing, and
#     30-89 DPD, each as a % of that segment's END-OF-PERIOD loan balance.
#     Sources are RC-N column C (nonaccrual), column B (90+), column A (30-89).
#
#   * FLOW rate: YTD net charge-offs, ANNUALIZED, over the segment's 5-quarter
#     average balance (CAVG5) -- i.e. UBPR "PCTOFANN":
#         rate = NCO_ytd * (4 / quarter) / CAVG5 * 100
#     Emitted ONLY where a CAVG5 average-balance field exists (cavg5 is None
#     otherwise; we never substitute an EOP denominator for a flow measure, so
#     no segment NCO rate is shown for segments lacking CAVG5 -- their $ values
#     are still shown). This mirrors the published total NTLNLSR exactly.
#
#   * Dollar balances ($000s) for each of the four problem buckets.
#
# Generating the metric names, categories, dollar/inverse membership, metric
# definitions, the FDIC field projection (FF), AND the calculator from this one
# table guarantees none of them can ever drift apart.
#
# Every field below is confirmed present in risview_properties.yaml (the FDIC
# financials field dictionary). Note: several segment splits have limited
# history -- the owner-occ / non-owner-occ NFNR split begins 2007Q1, and the
# auto loan breakout begins 2011Q1; before those dates the source fields are
# null and the affected metrics correctly read N/A (never fabricated).
#
#   (label, loan_balance, nonaccrual, p90, p30, nco_ytd, cavg5_or_None)
# =============================================================================
ASSET_QUALITY_SEGMENTS = [
    ("Real Estate",            "LNRE",     "NARE",     "P9RE",     "P3RE",     "NTRE",     "LNRE5"),
    ("RE 1-4 Family",          "LNRERES",  "NARERES",  "P9RERES",  "P3RERES",  "NTRERES",  "LNRERES5"),
    ("RE Multifamily",         "LNREMULT", "NAREMULT", "P9REMULT", "P3REMULT", "NTREMULT", None),
    ("RE Construction",        "LNRECONS", "NARECONS", "P9RECONS", "P3RECONS", "NTRECONS", None),
    ("RE Farmland",            "LNREAG",   "NAREAG",   "P9REAG",   "P3REAG",   "NTREAG",   None),
    ("RE Commercial (NFNR)",   "LNRENRES", "NARENRES", "P9RENRES", "P3RENRES", "NTRENRES", None),
    ("RE NFNR Owner-Occ",      "LNRENROW", "NARENROW", "P9RENROW", "P3RENROW", "NTRENROW", None),
    ("RE NFNR Non-Owner-Occ",  "LNRENROT", "NARENROT", "P9RENROT", "P3RENROT", "NTRENROT", None),
    ("C&I",                    "LNCI",     "NACI",     "P9CI",     "P3CI",     "NTCI",     "LNCI5"),
    ("Credit Cards",           "LNCRCD",   "NACRCD",   "P9CRCD",   "P3CRCD",   "NTCRCD",   "LNCRCD5"),
    ("Auto",                   "LNAUTO",   "NAAUTO",   "P9AUTO",   "P3AUTO",   "NTAUTO",   "LNAUTO5"),
    ("Consumer (Total)",       "LNCON",    "NACON",    "P9CON",    "P3CON",    "NTCON",    "LNCON5"),
    ("Agriculture",            "LNAG",     "NAAG",     "P9AG",     "P3AG",     "NTAG",     "LNAG5"),
    ("Leases",                 "LS",       "NALS",     "P9LS",     "P3LS",     "NTLS",     None),
    ("Other Loans",            "LNOTHER",  "NAOTHLN",  "P9OTHLN",  "P3OTHLN",  "NTOTHER",  None),
]

SEG_LABELS = [s[0] for s in ASSET_QUALITY_SEGMENTS]
# Only segments with a real CAVG5 average-balance field get a flow NCO rate.
SEG_NCO_RATE_LABELS = [s[0] for s in ASSET_QUALITY_SEGMENTS if s[6] is not None]


# --- Metric-name builders: the ONLY place these strings are defined. ---------
def _nm_na_rate(seg):  return f"Nonaccrual / {seg}"
def _nm_p9_rate(seg):  return f"90+ DPD / {seg}"
def _nm_p3_rate(seg):  return f"30-89 DPD / {seg}"
def _nm_nco_rate(seg): return f"Net Charge-Offs / {seg}"
def _nm_na_usd(seg):   return f"Nonaccrual ($) \u2014 {seg}"
def _nm_p9_usd(seg):   return f"90+ DPD ($) \u2014 {seg}"
def _nm_p3_usd(seg):   return f"30-89 DPD ($) \u2014 {seg}"
def _nm_nco_usd(seg):  return f"Net Charge-Offs YTD ($) \u2014 {seg}"


GROSS_RECOVERY_RATE_METRIC = "Gross Recovery Rate (YTD)"

NCO_SEG_RATE_METRICS = [_nm_nco_rate(s) for s in SEG_NCO_RATE_LABELS]
NCO_SEG_USD_METRICS = [_nm_nco_usd(s) for s in SEG_LABELS]
NCO_SEG_METRICS = NCO_SEG_RATE_METRICS + [GROSS_RECOVERY_RATE_METRIC] + NCO_SEG_USD_METRICS

NA_SEG_RATE_METRICS = [_nm_na_rate(s) for s in SEG_LABELS]
NA_SEG_USD_METRICS = [_nm_na_usd(s) for s in SEG_LABELS]
NA_SEG_METRICS = NA_SEG_RATE_METRICS + NA_SEG_USD_METRICS

P9_SEG_RATE_METRICS = [_nm_p9_rate(s) for s in SEG_LABELS]
P9_SEG_USD_METRICS = [_nm_p9_usd(s) for s in SEG_LABELS]
P9_SEG_METRICS = P9_SEG_RATE_METRICS + P9_SEG_USD_METRICS

P3_SEG_RATE_METRICS = [_nm_p3_rate(s) for s in SEG_LABELS]
P3_SEG_USD_METRICS = [_nm_p3_usd(s) for s in SEG_LABELS]
P3_SEG_METRICS = P3_SEG_RATE_METRICS + P3_SEG_USD_METRICS

ALL_SEG_RATE_METRICS = (NCO_SEG_RATE_METRICS + NA_SEG_RATE_METRICS
                        + P9_SEG_RATE_METRICS + P3_SEG_RATE_METRICS)
ALL_SEG_USD_METRICS = (NCO_SEG_USD_METRICS + NA_SEG_USD_METRICS
                       + P9_SEG_USD_METRICS + P3_SEG_USD_METRICS)

DOLLAR_METRICS = {
    'Total Assets', 'Total Deposits', 'Gross Loans & Leases', 'Net Loans & Leases',
    'Total Securities', 'Total Earning Assets', 'Total Equity Capital', 'Tier 1 Capital',
    'Risk-Weighted Assets', 'Net Income (YTD)', 'Net Income (Quarter)',
    'Allowance for Credit Losses', 'Noncurrent Loans',
    'Gross Charge-Offs (YTD)', 'Gross Charge-Offs (Quarter)',
    'Gross Recoveries (YTD)', 'Gross Recoveries (Quarter)',
    'Brokered Deposits',
} | set(ALL_SEG_USD_METRICS)

# -----------------------------------------------------------------------------
# Categories now hold EXPLICIT metric lists (previously index slices of
# METRIC_ORDER, which silently mis-aligned whenever a metric was added). The
# canonical METRIC_ORDER is now DERIVED by flattening the categories, so the two
# can never disagree.
# -----------------------------------------------------------------------------
METRIC_CATEGORIES = [
    ("Earnings & Profitability", [
        'Return on Assets', 'Quarterly Return on Assets', 'Pretax Return on Assets',
        'Return on Equity', 'Quarterly Return on Equity',
        'Net Operating Income to Assets', 'Interest Income to Average Assets',
        'Interest Expense to Average Assets', 'Pre-Provision Net Revenue to Average Assets',
        'Provision for Credit Losses to Average Assets',
    ]),
    ("Efficiency & Margin", [
        'Yield on Earning Assets', 'Net Interest Margin',
        'Cost of Funding Earning Assets (YTD)', 'Cost of Funding Earning Assets (Quarterly)',
        'Earning Assets / Total Assets', 'Efficiency Ratio',
        'Noninterest Expense to Average Assets', 'Salaries and Benefits to Average Assets',
        'Noninterest Income to Average Assets',
    ]),
    ("Capitalization", [
        'Common Equity Tier 1 (CET1) Ratio', 'Tier 1 Risk-Based Capital Ratio',
        'Leverage (Core Capital) Ratio', 'Total Risk-Based Capital Ratio',
    ]),
    ("Asset Quality", [
        'Net Charge-Offs / Total Loans & Leases', 'ACL / Total Loans & Leases',
        'ACL / Nonaccrual Loans', 'ACL / 90+ DPD & Nonaccrual',
        'Loan Loss Reserve / Noncurrent Loans', 'Nonaccrual & OREO / Total Loans & OREO',
        '30-89 DPD / Total Loans', '90+ DPD / Total Loans',
        'Nonaccrual / Total Loans', '90+ DPD & Nonaccrual / Total Loans',
    ]),
    ("Loan & Lease", [
        'Net Loan Growth Rate', 'Earnings Coverage of Net Loan Charge-Offs',
        'Loan and Lease Loss Provision to Net Charge-Offs', 'Net Charge-Offs / ACL',
        'Net Loans and Leases to Assets',
    ]),
    ("Funding & Liquidity", [
        'Net Loans and Leases to Deposits', 'Core Deposits to Total Deposits',
        'Noninterest-Bearing Deposits to Total Deposits', 'Brokered Deposits to Total Deposits',
        'Volatile Liabilities to Total Assets',
    ]),
    ("Credit Concentration", [
        'Real Estate Loans to Tier 1 + ACL',
        'RE Construction and Land Development to Tier 1 + ACL',
        '1-4 Family Construction to Tier 1 + ACL',
        'Other Construction & Land Dev to Tier 1 + ACL',
        'Secured by Farmland to Tier 1 + ACL', '1-4 Family Residential to Tier 1 + ACL',
        'Revolving Home Equity to Tier 1 + ACL', 'Closed-End 1st Lien to Tier 1 + ACL',
        'Closed-End Jr Lien to Tier 1 + ACL', 'Multifamily RE to Tier 1 + ACL',
        'Non-Farm Non-Residential RE to Tier 1 + ACL', 'NFNR: Owner Occupied to Tier 1 + ACL',
        'NFNR: Non-Owner Occupied to Tier 1 + ACL', 'Commercial RE to Tier 1 + ACL',
        'Non-Owner Occupied CRE to Tier 1 + ACL',
        'Non-Owner Occupied CRE 3-Year Growth Rate', 'C&I Loans to Tier 1 + ACL',
        'Loans to Individuals to Tier 1 + ACL', 'Credit Cards to Tier 1 + ACL',
        'Auto Loans to Tier 1 + ACL', 'Agriculture Loans to Tier 1 + ACL',
        'Loans to NDFIs and Other to Tier 1 + ACL',
    ]),
    ("Growth", [
        'Total Asset Growth Rate', 'Tier 1 Capital Growth Rate',
    ]),
    # ---- Segment-level asset-quality detail (generated above) ---------------
    ("Net Charge-Offs by Segment", NCO_SEG_METRICS),
    ("Nonaccrual by Segment", NA_SEG_METRICS),
    ("90+ Day Past Due by Segment", P9_SEG_METRICS),
    ("30-89 Day Past Due by Segment", P3_SEG_METRICS),
    ("Key Financials", [
        'Total Assets', 'Total Deposits', 'Gross Loans & Leases', 'Net Loans & Leases',
        'Total Securities', 'Total Earning Assets', 'Total Equity Capital', 'Tier 1 Capital',
        'Risk-Weighted Assets', 'Net Income (YTD)', 'Net Income (Quarter)',
        'Allowance for Credit Losses', 'Gross Charge-Offs (YTD)',
        'Gross Charge-Offs (Quarter)', 'Gross Recoveries (YTD)',
        'Gross Recoveries (Quarter)', 'Noncurrent Loans', 'Brokered Deposits',
    ]),
]

# Canonical ordering derived from the categories (flatten, order preserved).
METRIC_ORDER = [m for _cat, _ms in METRIC_CATEGORIES for m in _ms]

METRIC_TO_CATEGORY = {}
for cat_name, cat_metrics in METRIC_CATEGORIES:
    for m in cat_metrics:
        METRIC_TO_CATEGORY[m] = cat_name

# Concentration metrics share the (CECL-adjusted) Tier 1 + ACL denominator.
# Referenced BY NAME so it is robust to category reordering.
CONCENTRATION_METRICS = list(next(ms for name, ms in METRIC_CATEGORIES
                                  if name == "Credit Concentration"))

CATEGORY_ACCENTS = {
    "Key Financials": "#0f172a",
    "Earnings & Profitability": "#16a34a", "Efficiency & Margin": "#2563eb",
    "Capitalization": "#d97706", "Asset Quality": "#dc2626",
    "Loan & Lease": "#7c3aed", "Funding & Liquidity": "#0891b2",
    "Credit Concentration": "#64748b", "Growth": "#4f46e5",
    # Segment asset-quality categories, shaded as a delinquency "heat" ramp:
    # 30-89 (earliest) -> 90+ -> nonaccrual -> charge-off (loss realized).
    "30-89 Day Past Due by Segment": "#d97706",
    "90+ Day Past Due by Segment": "#ea580c",
    "Nonaccrual by Segment": "#dc2626",
    "Net Charge-Offs by Segment": "#991b1b",
}
CATEGORY_BG = {
    "Key Financials": "#f1f5f9",
    "Earnings & Profitability": "#f0fdf4", "Efficiency & Margin": "#eff6ff",
    "Capitalization": "#fffbeb", "Asset Quality": "#fef2f2",
    "Loan & Lease": "#f5f3ff", "Funding & Liquidity": "#ecfeff",
    "Credit Concentration": "#f8fafc", "Growth": "#eef2ff",
    "30-89 Day Past Due by Segment": "#fffbeb",
    "90+ Day Past Due by Segment": "#fff7ed",
    "Nonaccrual by Segment": "#fff1f2",
    "Net Charge-Offs by Segment": "#fef2f2",
}

CATEGORY_SHORT_LABELS = {
    "Earnings & Profitability": "Earnings & Profitability",
    "Efficiency & Margin": "Efficiency & Margin",
    "Capitalization": "Capitalization",
    "Asset Quality": "Asset Quality",
    "Loan & Lease": "Loan & Lease",
    "Funding & Liquidity": "Funding & Liquidity",
    "Credit Concentration": "Credit Concentration",
    "Growth": "Growth",
    "Key Financials": "Key Financials",
    "Net Charge-Offs by Segment": "NCO by Segment",
    "Nonaccrual by Segment": "Nonaccrual by Segment",
    "90+ Day Past Due by Segment": "90+ DPD by Segment",
    "30-89 Day Past Due by Segment": "30-89 DPD by Segment",
}

INVERSE_METRICS = {
    'Efficiency Ratio', 'Interest Expense to Average Assets',
    'Cost of Funding Earning Assets (YTD)', 'Cost of Funding Earning Assets (Quarterly)',
    'Noninterest Expense to Average Assets',
    'Salaries and Benefits to Average Assets',
    'Provision for Credit Losses to Average Assets',
    'Net Charge-Offs / Total Loans & Leases',
    'Nonaccrual & OREO / Total Loans & OREO',
    '30-89 DPD / Total Loans', '90+ DPD / Total Loans',
    'Nonaccrual / Total Loans', '90+ DPD & Nonaccrual / Total Loans',
    'Net Charge-Offs / ACL',
    'Brokered Deposits to Total Deposits', 'Volatile Liabilities to Total Assets',
    'Gross Charge-Offs (YTD)', 'Gross Charge-Offs (Quarter)', 'Noncurrent Loans',
# Every segment delinquency/nonaccrual/charge-off rate AND dollar balance is a
# "higher = worse" measure (consistent with how the total-level problem-loan
# dollar metrics above are already treated). The recovery RATE is deliberately
# excluded -- a higher recovery rate is favorable.
} | set(ALL_SEG_RATE_METRICS) | set(ALL_SEG_USD_METRICS)

NON_PERCENT_RATIO_METRICS = {
    'Earnings Coverage of Net Loan Charge-Offs',
    'ACL / Nonaccrual Loans',
    'ACL / 90+ DPD & Nonaccrual',
}

METRIC_DEFINITIONS = {
    'Total Assets': "Key Financials \xb7 Sum of all assets. FDIC field: ASSET. Call Report Schedule RC line 12. Values in $000s.",
    'Total Deposits': "Key Financials \xb7 Total domestic and foreign deposits. FDIC field: DEP. Schedule RC line 13. Values in $000s.",
    'Gross Loans & Leases': "Key Financials \xb7 Total loans and lease financing receivables before ACL. FDIC field: LNLSGR. Schedule RC-C. Values in $000s.",
    'Net Loans & Leases': "Key Financials \xb7 Total loans and leases NET of allowance for credit losses. FDIC field: LNLSNET. Used in the UBPR Net Loans & Leases / Deposits ratio. Values in $000s.",
    'Total Securities': "Key Financials \xb7 HTM + AFS securities. FDIC field: SC. Schedule RC line 2. Values in $000s.",
    'Total Earning Assets': "Key Financials \xb7 Interest-bearing balances + securities + net loans + fed funds sold + trading assets. FDIC field: ERNAST. Values in $000s.",
    'Total Equity Capital': "Key Financials \xb7 Total equity capital including AOCI. FDIC field: EQ. Schedule RC line 28. Values in $000s.",
    'Tier 1 Capital': "Key Financials \xb7 Regulatory Tier 1 capital (CET1 + AT1). FDIC field: RBCT1J. Schedule RC-R. Values in $000s.",
    'Risk-Weighted Assets': "Key Financials \xb7 Total risk-weighted assets. FDIC field: RWAJ. Schedule RC-R. Values in $000s.",
    'Net Income (YTD)': "Key Financials \xb7 Year-to-date net income after taxes. FDIC field: NETINC. Values in $000s.",
    'Net Income (Quarter)': "Key Financials \xb7 Current-quarter net income after taxes. FDIC field: NETINCQ. Values in $000s.",
    'Allowance for Credit Losses': "Key Financials \xb7 ACL on loans & leases HFI (CECL). FDIC field: LNATRES. RC line 4.c. Values in $000s.",
    'Gross Charge-Offs (YTD)': "Key Financials \xb7 Year-to-date gross charge-offs. FDIC field: DRLNLS. Values in $000s.",
    'Gross Charge-Offs (Quarter)': "Key Financials \xb7 Current-quarter gross charge-offs. FDIC field: DRLNLSQ. Values in $000s.",
    'Gross Recoveries (YTD)': "Key Financials \xb7 Year-to-date gross recoveries. FDIC field: CRLNLS. Values in $000s.",
    'Gross Recoveries (Quarter)': "Key Financials \xb7 Current-quarter recoveries. FDIC field: CRLNLSQ. Values in $000s.",
    'Noncurrent Loans': "Key Financials \xb7 Nonaccrual + 90+ DPD still accruing. Computed: NALNLS + P9LNLS. Values in $000s.",
    'Brokered Deposits': "Key Financials \xb7 All brokered deposits. FDIC field: BRO. Schedule RC-E. Values in $000s.",
    'Return on Assets': "Earnings \xb7 Annualized net income as % of average total assets. FDIC field: ROA. Already annualized by FDIC.",
    'Quarterly Return on Assets': "Earnings \xb7 Current-quarter net income as % of average total assets. FDIC field: ROAQ.",
    'Pretax Return on Assets': "Earnings \xb7 Pretax net operating income (TE) as % of avg assets. FDIC field: ROAPTX. UBPR Pg1 #12 (UBPRE009).",
    'Return on Equity': "Earnings \xb7 Annualized net income as % of avg equity. FDIC field: ROE. Already annualized by FDIC.",
    'Quarterly Return on Equity': "Earnings \xb7 Current-quarter net income as % of average equity. FDIC field: ROEQ.",
    'Net Operating Income to Assets': "Earnings \xb7 Adjusted net operating income as % of avg assets. FDIC field: NOIJY. UBPR Pg1 #13 (UBPRE010).",
    'Interest Income to Average Assets': "Earnings \xb7 Total interest income (TE) as % of avg assets. FDIC field: INTINCR. UBPR Pg1 #1 (UBPRE001).",
    'Interest Expense to Average Assets': "Earnings \xb7 Total interest expense as % of avg assets. FDIC field: EINTEXPR. UBPR Pg1 #2 (UBPRE002).",
    'Pre-Provision Net Revenue to Average Assets': "Earnings \xb7 PPNR = (II \u2212 IE + NII \u2212 NIE) / Avg Assets. UBPR Pg1 #6 (UBPRPG69). Computed from component fields.",
    'Provision for Credit Losses to Average Assets': "Earnings \xb7 Provision for credit losses as % of avg assets. FDIC field: ELNATRR. UBPR Pg1 #7 (UBPRE006).",
    'Yield on Earning Assets': "Margin \xb7 Interest income (TE, annualized) as % of avg earning assets. FDIC field: INTINCY. UBPR Pg1 #19 (UBPRE016).",
    'Net Interest Margin': "Margin \xb7 Net interest income (TE) as % of avg earning assets. Typical: 2.50\u20133.80%. FDIC field: NIMY. UBPR Pg1 #21 (UBPRE018).",
    'Cost of Funding Earning Assets (YTD)': "Margin \xb7 YTD/annualized interest expense as % of avg earning assets. FDIC field: INTEXPY. Source formula: EINTEXPA / ERNAST5. This is the YTD cost-of-funding companion to INTEXPYQ.",
    'Cost of Funding Earning Assets (Quarterly)': "Margin \xb7 Current-quarter interest expense as % of avg earning assets. FDIC field: INTEXPYQ. UBPR Pg1 #20 (UBPRE017).",
    'Earning Assets / Total Assets': "Margin \xb7 Earning assets as % of total assets. FDIC field: ERNASTR. UBPR Pg1 #17 (UBPRE014).",
    'Efficiency Ratio': "Margin \xb7 NIE / (NII + noninterest income). Lower ratio indicates greater efficiency. FDIC field: EEFFR. UBPR Pg3 (UBPRE095).",
    'Noninterest Expense to Average Assets': "Margin \xb7 Total noninterest expense as % of avg assets. FDIC field: NONIXR. UBPR Pg1 #5 (UBPRE005).",
    'Salaries and Benefits to Average Assets': "Margin \xb7 Personnel expense as % of avg assets. FDIC field: ESALR. UBPR Pg3.",
    'Noninterest Income to Average Assets': "Margin \xb7 Total fee/noninterest income as % of avg assets. FDIC field: NONIIR. UBPR Pg1 #4 (UBPRE004).",
    'Common Equity Tier 1 (CET1) Ratio': "Capitalization \xb7 CET1 capital to RWA. Well-capitalized: \u22656.5%. FDIC field: IDT1CER. UBPR Pg11. (Regulatory ratio: intentionally reflects the bank's elected CECL transition relief; NOT CECL-adjusted here.)",
    'Tier 1 Risk-Based Capital Ratio': "Capitalization \xb7 Tier 1 capital to RWA. Well-capitalized: \u22658%. FDIC field: IDT1RWAJR. UBPR Pg11. (Regulatory ratio: reflects elected CECL transition relief; NOT CECL-adjusted here.)",
    'Leverage (Core Capital) Ratio': "Capitalization \xb7 Tier 1 capital to avg total assets. Well-capitalized: \u22655%. FDIC field: RBC1AAJ. UBPR Pg1 #33 (UBPRD486). (Regulatory ratio: reflects elected CECL transition relief; NOT CECL-adjusted here.)",
    'Total Risk-Based Capital Ratio': "Capitalization \xb7 Total capital (T1 + T2) to RWA. Well-capitalized: \u226510%. FDIC field: RBCRWAJ. UBPR Pg1 #34 (UBPRD488). (Regulatory ratio: reflects elected CECL transition relief; NOT CECL-adjusted here.)",
    'Net Charge-Offs / Total Loans & Leases': "Asset Quality \xb7 Net charge-offs as % of avg total L&L, annualized. FDIC field: NTLNLSR. UBPR Pg1 #22 (UBPRE019).",
    'ACL / Total Loans & Leases': "Asset Quality \xb7 ACL as % of LN&LS HFI. FDIC field: LNATRESR. UBPR Pg1 #24 (UBPRE022).",
    'ACL / Nonaccrual Loans': "Asset Quality \xb7 ACL as a MULTIPLE of nonaccrual loans. Well-reserved: \u22651.00x. UBPR Pg1 #26 (UBPRE395). Shown as X multiplier (1.51x = 151%).",
    'ACL / 90+ DPD & Nonaccrual': "Asset Quality \xb7 ACL as a MULTIPLE of (nonaccrual + 90+ DPD still accruing). Shown as X multiplier. Dashboard-computed.",
    'Loan Loss Reserve / Noncurrent Loans': "Asset Quality \xb7 ACL as % of noncurrent loans. Below 100% = reserves may not cover problems. FDIC field: LNRESNCR. UBPR Pg1 #36 (UBPRNC98).",
    'Nonaccrual & OREO / Total Loans & OREO': "Asset Quality \xb7 (Nonaccrual + OREO + 90+ DPD) / (total loans + OREO). Broadest NPA measure. UBPR Pg1 #29 (UBPRE549).",
    '30-89 DPD / Total Loans': "Asset Quality \xb7 Early-stage delinquency. UBPR Pg1 #27 (UBPRE544).",
    '90+ DPD / Total Loans': "Asset Quality \xb7 Seriously delinquent but still accruing. UBPR Pg8. Dashboard-computed.",
    'Nonaccrual / Total Loans': "Asset Quality \xb7 Loans where interest recognition suspended. UBPR Pg8. Dashboard-computed.",
    '90+ DPD & Nonaccrual / Total Loans': "Asset Quality \xb7 Total noncurrent as % of portfolio. UBPR Pg1 #28 (UBPR7414).",
    'Net Loan Growth Rate': "Loan & Lease \xb7 YoY growth of net L&L. >20% YoY triggers regulatory scrutiny. UBPR Pg1 #39 (UBPRE027).",
    'Earnings Coverage of Net Loan Charge-Offs': "Loan & Lease \xb7 Times net income covers NCOs. <1.0x means losing money after credit losses. FDIC field: IDERNCVR. UBPR Pg1 #23 (UBPRE020). Shown as X multiplier.",
    'Loan and Lease Loss Provision to Net Charge-Offs': "Loan & Lease \xb7 Provision expense as % of NCOs. >100% = building reserves. FDIC field: ELNANTR. UBPR Pg7.",
    'Net Charge-Offs / ACL': "Loan & Lease \xb7 Rolling 4-quarter NCOs as % of ACL. Dashboard-computed; requires four contiguous quarters, else N/A.",
    'Net Loans and Leases to Assets': "Loan & Lease \xb7 Net L&L as % of total assets. FDIC field: LNLSNTV. UBPR Pg1 #31 (UBPRE024).",
    'Net Loans and Leases to Deposits': "Funding \xb7 Net loans as % of total deposits. Typical: 70\u201395%. FDIC field: LNLSDEPR. UBPR Pg1 #32 (UBPRE600).",
    'Core Deposits to Total Deposits': "Funding \xb7 Core deposits as % of total deposits. Computed: COREDEP / DEP. UBPR Pg4.",
    'Noninterest-Bearing Deposits to Total Deposits': "Funding \xb7 Domestic NIB deposits as % of total deposits. Computed: DEPNIDOM / DEP. UBPR Pg4.",
    'Brokered Deposits to Total Deposits': "Funding \xb7 All brokered deposits as % of total deposits. \u226520% signals vulnerability. Computed: BRO / DEP. UBPR Pg4.",
    'Volatile Liabilities to Total Assets': "Funding \xb7 Volatile liabilities as % of total assets. FDIC field: VOLIABR. UBPR Pg10.",
    'Real Estate Loans to Tier 1 + ACL': "Concentration \xb7 Total RE loans as % of Tier 1 + ACL. UBPR Pg7B 7B.1 (UBPRE884). Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'RE Construction and Land Development to Tier 1 + ACL': "Concentration \xb7 Construction & land dev as % of Tier 1 + ACL. UBPR Pg7B 7B.2 (UBPRD490). Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    '1-4 Family Construction to Tier 1 + ACL': "Concentration \xb7 1-4 family construction as % of Tier 1 + ACL. UBPR Pg7B 7B.3 (UBPRE632). Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Other Construction & Land Dev to Tier 1 + ACL': "Concentration \xb7 Other construction & land dev as % of Tier 1 + ACL. UBPR Pg7B 7B.4. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Secured by Farmland to Tier 1 + ACL': "Concentration \xb7 Farmland RE as % of Tier 1 + ACL. UBPR Pg7B 7B.5. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    '1-4 Family Residential to Tier 1 + ACL': "Concentration \xb7 Total 1-4 family residential as % of Tier 1 + ACL. UBPR Pg7B 7B.6. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Revolving Home Equity to Tier 1 + ACL': "Concentration \xb7 HELOCs as % of Tier 1 + ACL. UBPR Pg7B 7B.7. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Closed-End 1st Lien to Tier 1 + ACL': "Concentration \xb7 Closed-end first lien 1-4 family as % of Tier 1 + ACL. UBPR Pg7B 7B.8. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Closed-End Jr Lien to Tier 1 + ACL': "Concentration \xb7 Junior lien as % of Tier 1 + ACL. UBPR Pg7B 7B.9. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Multifamily RE to Tier 1 + ACL': "Concentration \xb7 Multifamily (5+ units) as % of Tier 1 + ACL. UBPR Pg7B 7B.10. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Non-Farm Non-Residential RE to Tier 1 + ACL': "Concentration \xb7 NFNR total as % of Tier 1 + ACL. UBPR Pg7B 7B.11. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'NFNR: Owner Occupied to Tier 1 + ACL': "Concentration \xb7 Owner-occupied NFNR as % of Tier 1 + ACL. UBPR Pg7B 7B.12. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'NFNR: Non-Owner Occupied to Tier 1 + ACL': "Concentration \xb7 Non-owner-occupied NFNR (investor CRE) as % of Tier 1 + ACL. UBPR Pg7B 7B.13. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Commercial RE to Tier 1 + ACL': "Concentration \xb7 UBPR Total CRE = Construction + Multifamily + NFNR total + LNCOMRE, as % of Tier 1 + ACL. UBPR Pg7B 7B.26. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Non-Owner Occupied CRE to Tier 1 + ACL': "Concentration \xb7 NOO CRE = Construction + Multifamily + NFNR NOO + LNCOMRE, as % of Tier 1 + ACL. Interagency CRE guidance metric. UBPR Pg7B 7B.24. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Non-Owner Occupied CRE 3-Year Growth Rate': "Growth \xb7 3-year growth of NOO CRE. >300% concentration AND >36% 3-year growth trigger enhanced risk management. UBPR Pg7B 7B.25. (Dollar-balance growth: NOT affected by the CECL denominator adjustment.)",
    'C&I Loans to Tier 1 + ACL': "Concentration \xb7 C&I loans as % of Tier 1 + ACL. UBPR Pg7B 7B.17 (UBPRE887). Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Loans to Individuals to Tier 1 + ACL': "Concentration \xb7 Total consumer loans as % of Tier 1 + ACL. UBPR Pg7B 7B.18 (UBPRE888). Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Credit Cards to Tier 1 + ACL': "Concentration \xb7 Credit card loans as % of Tier 1 + ACL. UBPR Pg7B 7B.19. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Auto Loans to Tier 1 + ACL': "Concentration \xb7 Auto loans as % of Tier 1 + ACL. UBPR Pg7B 7B.20. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Agriculture Loans to Tier 1 + ACL': "Concentration \xb7 Agriculture loans (non-RE) as % of Tier 1 + ACL. UBPR Pg7B 7B.16 (UBPRE886). Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Loans to NDFIs and Other to Tier 1 + ACL': "Concentration \xb7 Loans to nondepository FIs and other as % of Tier 1 + ACL. UBPR Pg7B 7B.22. Denominator nets out the CECL transitional add-back during 2020\u20132024 (OCC Bull. 2020-90; Fed SR 20-8), matching UBPR.",
    'Total Asset Growth Rate': "Growth \xb7 YoY growth of total assets. \u226530% YoY invites regulatory questions. UBPR Pg1 #37 (UBPR7316).",
    'Tier 1 Capital Growth Rate': "Growth \xb7 YoY growth of Tier 1 capital. UBPR Pg1 #38 (UBPR7408).",
}


def _build_segment_definitions():
    """Generate METRIC_DEFINITIONS entries for every segment metric directly
    from ASSET_QUALITY_SEGMENTS, so each one documents its exact FDIC source
    field(s) and methodology -- and stays in lockstep with the calculator."""
    d = {}
    hist = ("Limited history for some splits: the NFNR owner-occ / non-owner-occ "
            "breakout begins 2007Q1 and the auto-loan breakout begins 2011Q1; "
            "before those dates the source field is null and the metric reads N/A.")
    for (label, bal, na, p9, p3, nt, cavg5) in ASSET_QUALITY_SEGMENTS:
        d[_nm_na_rate(label)] = (
            f"Nonaccrual by Segment \xb7 Nonaccrual {label} loans as % of {label} loans (end-of-period). "
            f"Computed: {na} / {bal} \u00d7 100. Point-in-time stock ratio (RC-N col C); higher = weaker. {hist}")
        d[_nm_p9_rate(label)] = (
            f"90+ Day Past Due by Segment \xb7 {label} loans 90+ days past due and still accruing, as % of "
            f"{label} loans (EOP). Computed: {p9} / {bal} \u00d7 100. Point-in-time (RC-N col B). {hist}")
        d[_nm_p3_rate(label)] = (
            f"30-89 Day Past Due by Segment \xb7 {label} loans 30-89 days past due, as % of {label} loans "
            f"(EOP). Computed: {p3} / {bal} \u00d7 100. Early-stage delinquency (RC-N col A); typically "
            f"migrates to nonaccrual within 1-2 quarters. {hist}")
        d[_nm_na_usd(label)] = (
            f"Nonaccrual by Segment \xb7 Nonaccrual {label} loans. FDIC field: {na}. Values in $000s.")
        d[_nm_p9_usd(label)] = (
            f"90+ Day Past Due by Segment \xb7 {label} loans 90+ DPD and still accruing. FDIC field: {p9}. Values in $000s.")
        d[_nm_p3_usd(label)] = (
            f"30-89 Day Past Due by Segment \xb7 {label} loans 30-89 DPD. FDIC field: {p3}. Values in $000s.")
        d[_nm_nco_usd(label)] = (
            f"Net Charge-Offs by Segment \xb7 Year-to-date net charge-offs on {label} loans. FDIC field: {nt}. "
            f"Values in $000s (YTD basis; resets each Q1).")
        if cavg5 is not None:
            d[_nm_nco_rate(label)] = (
                f"Net Charge-Offs by Segment \xb7 YTD net charge-offs on {label} loans, ANNUALIZED, as % of "
                f"average {label} loans. Computed: {nt} \u00d7 (4/quarter) / {cavg5} (CAVG5 5-quarter average) "
                f"\u00d7 100. UBPR PCTOFANN methodology, consistent with the published total (NTLNLSR). Higher = worse.")
    d[GROSS_RECOVERY_RATE_METRIC] = (
        "Net Charge-Offs by Segment \xb7 YTD gross recoveries as % of YTD gross charge-offs. "
        "Computed: CRLNLS / DRLNLS \u00d7 100. Higher = more charged-off principal recovered (favorable).")
    return d


METRIC_DEFINITIONS.update(_build_segment_definitions())

# -----------------------------------------------------------------------------
# FDIC financials field projection (FF).
# Built at MODULE level: the original base projection PLUS every distinct field
# the segment table needs (nonaccrual / 90+ / 30-89 / net charge-off detail and
# the CAVG5 average-balance denominators), de-duplicated against the base. The
# class attribute BankDataRepository.FF is set to FULL_FF below. Still ONE FDIC
# call per bank -- the field list simply grows; no extra requests.
# -----------------------------------------------------------------------------
_BASE_FF = (
    "CERT,REPDTE,ASSET,DEP,BRO,LNLSGR,LNLSNET,SC,ERNAST,RWAJ,"
    "LNRE,LNRECONS,LNRECNFM,LNRECNOT,LNREAG,LNRERES,LNRELOC,LNRERSFM,LNRERSF2,"
    "LNREMULT,LNRENRES,LNRENROW,LNRENROT,LNCOMRE,"
    "LNCI,LNAG,LNCON,LNCRCD,LNAUTO,LNCONOTH,LNOTHER,"
    "LNATRES,NALNLS,OREOTH,P3LNLS,P9LNLS,RBCT1J,CT1BADJ,EQ,EQPP,DRLNLS,DRLNLSQ,"
    "CRLNLS,CRLNLSQ,NTLNLSQ,NETINC,NETINCQ,ERNASTR,NIMY,NTLNLSR,LNATRESR,ROA,ROAQ,"
    "ROE,ROEQ,RBC1AAJ,RBCRWAJ,LNLSDEPR,LNLSNTV,"
    "EEFFR,ELNANTR,IDERNCVR,IDT1CER,IDT1RWAJR,INTEXPY,INTEXPYQ,NONIIR,COREDEP,ROAPTX,"
    "NONIXR,DEPNIDOM,LNRESNCR,VOLIABR,NOIJY,ESALR,INTINCR,EINTEXPR,ELNATRR,INTINCY,"
    "RBCEQUP,EQUPGR"  # CECL concentration add-back (RC-R ln2 - RC 26.a)
)

# Every distinct field referenced by the segment table (loan-balance denominators,
# nonaccrual/90+/30-89 detail, segment net charge-offs, CAVG5 averages), plus the
# total-level recovery numerator used by Gross Recovery Rate.
_AQ_SEG_FIELDS = set()
for _seg in ASSET_QUALITY_SEGMENTS:
    _AQ_SEG_FIELDS.update(f for f in (_seg[1], _seg[2], _seg[3], _seg[4], _seg[5]) if f)
    if _seg[6]:
        _AQ_SEG_FIELDS.add(_seg[6])
_AQ_SEG_FIELDS.update({"DRLNLS", "CRLNLS"})

_BASE_FF_SET = set(_BASE_FF.split(","))
_SEG_FF_EXTRA = sorted(f for f in _AQ_SEG_FIELDS if f not in _BASE_FF_SET)
FULL_FF = _BASE_FF + ("," + ",".join(_SEG_FF_EXTRA) if _SEG_FF_EXTRA else "")


def normalize_bank_name(n):
    if not n: return n
    if n in BANK_NAME_MAPPING: return BANK_NAME_MAPPING[n]
    u = n.upper().strip()
    for o, d in BANK_NAME_MAPPING.items():
        if o.upper().strip() == u: return d
    return n


def is_dollar_metric(m):
    return m in DOLLAR_METRICS


def is_inverse_metric(m):
    return m in INVERSE_METRICS


def is_multiplier_metric(m):
    return m in NON_PERCENT_RATIO_METRICS


def is_percent_metric(m):
    if m is None: return False
    if m in DOLLAR_METRICS: return False
    if m in NON_PERCENT_RATIO_METRICS: return False
    return True


def safe_div(numerator, denominator, scale=100.0):
    if numerator is None or denominator is None:
        return None
    if pd.isna(numerator) or pd.isna(denominator):
        return None
    try:
        n = float(numerator); d = float(denominator)
    except (TypeError, ValueError):
        return None
    if d == 0 or not np.isfinite(d) or not np.isfinite(n):
        return None
    return (n / d) * scale


def fmt_dollar(v):
    if v is None or pd.isna(v):
        return "N/A"
    try: v = float(v)
    except (TypeError, ValueError): return "N/A"
    av = abs(v); sign = "-" if v < 0 else ""
    if av >= 1_000_000: return f"{sign}${av / 1_000_000:,.1f}B"
    elif av >= 1_000:   return f"{sign}${av / 1_000:,.1f}M"
    elif av >= 1:       return f"{sign}${av:,.0f}K"
    else:               return "$0"


def fmt_val(v, m=None, with_unit=False):
    if v is None or pd.isna(v):
        return "N/A"
    if m and is_dollar_metric(m):
        return fmt_dollar(v)
    try:
        out = f"{float(v):.2f}"
    except (TypeError, ValueError):
        return "N/A"
    if with_unit and m:
        if is_percent_metric(m):
            return f"{out}%"
        if is_multiplier_metric(m):
            return f"{out}x"
    return out


def fmt_delta(curr, prev, metric=None):
    if curr is None or prev is None or pd.isna(curr) or pd.isna(prev):
        return ("\u2014", CS['neutral'])
    try:
        curr, prev = float(curr), float(prev)
    except (TypeError, ValueError):
        return ("\u2014", CS['neutral'])
    diff = curr - prev
    inverse = is_inverse_metric(metric) if metric else False
    if metric and is_dollar_metric(metric):
        if prev == 0: return ("\u2014", CS['neutral'])
        pct = (diff / abs(prev)) * 100
        display = f"{pct:+.1f}%"
        is_good = (pct > 0) if not inverse else (pct < 0)
    else:
        if metric and is_percent_metric(metric):
            display = f"{diff:+.2f} pp"
        elif metric and is_multiplier_metric(metric):
            display = f"{diff:+.2f}x"
        else:
            display = f"{diff:+.2f}"
        is_good = (diff > 0) if not inverse else (diff < 0)
    if abs(diff) < 1e-9:
        if metric and is_percent_metric(metric):
            zero_display = "0.00 pp"
        elif metric and is_multiplier_metric(metric):
            zero_display = "0.00x"
        else:
            zero_display = "0.00"
        return (zero_display, CS['neutral'])
    color = CS['good'] if is_good else CS['bad']
    return (display, color)


def calc_trend_change(start_value, end_value, metric=None):
    if start_value is None or end_value is None or pd.isna(start_value) or pd.isna(end_value):
        return np.nan
    try:
        sv, ev = float(start_value), float(end_value)
    except (TypeError, ValueError):
        return np.nan
    if not np.isfinite(sv) or not np.isfinite(ev):
        return np.nan
    if metric and is_dollar_metric(metric):
        return ((ev - sv) / abs(sv)) * 100 if sv != 0 else np.nan
    return ev - sv


def fmt_trend_change(value, metric=None):
    if value is None or pd.isna(value):
        return "N/A"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "N/A"
    if metric and is_dollar_metric(metric):
        return f"{v:+.2f}%"
    if metric and is_multiplier_metric(metric):
        return f"{v:+.2f}x"
    return f"{v:+.2f} pp"


def trend_direction_label(slope, eps=1e-9):
    if slope is None or pd.isna(slope):
        return "N/A"
    try:
        sl = float(slope)
    except (TypeError, ValueError):
        return "N/A"
    if abs(sl) <= eps:
        return "\u2192 Flat"
    return "\u2191 Up" if sl > 0 else "\u2193 Down"


# =============================================================================
# SVG MICRO-VISUALS (sparklines, percentile arcs) + period/rank helpers
# =============================================================================
def make_sparkline_svg(values, width=90, height=24, color=None, fill_color=None):
    color = color or CS['spark']
    fill_color = fill_color or CS['spark_area']
    clean = [float(v) for v in values if v is not None and not pd.isna(v)]
    if len(clean) < 2:
        return (f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" '
                f'xmlns="http://www.w3.org/2000/svg"><line x1="2" y1="{height/2:.1f}" '
                f'x2="{width-2:.1f}" y2="{height/2:.1f}" stroke="{CS["lighter"]}" '
                f'stroke-width="1" stroke-dasharray="2,2"/></svg>')
    vmin, vmax = min(clean), max(clean)
    rng = vmax - vmin if vmax != vmin else max(abs(vmax), 1) * 0.1
    pad = 3
    pts = []
    for i, v in enumerate(clean):
        x = (i / (len(clean) - 1)) * (width - 2) + 1
        y = (height - pad) - ((v - vmin) / rng) * (height - 2 * pad)
        pts.append((x, y))
    line_path = "M" + " L".join(f"{x:.1f},{y:.1f}" for x, y in pts)
    area_path = line_path + f" L{pts[-1][0]:.1f},{height - 0.5} L{pts[0][0]:.1f},{height - 0.5} Z"
    last_x, last_y = pts[-1]
    return (f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" '
            f'xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="none">'
            f'<path d="{area_path}" fill="{fill_color}" stroke="none"/>'
            f'<path d="{line_path}" fill="none" stroke="{color}" stroke-width="1.4" '
            f'stroke-linecap="round" stroke-linejoin="round"/>'
            f'<circle cx="{last_x:.1f}" cy="{last_y:.1f}" r="1.8" fill="{color}"/>'
            f'</svg>')


def svg_to_data_url(svg_str):
    encoded = base64.b64encode(svg_str.encode('utf-8')).decode('ascii')
    return f"data:image/svg+xml;base64,{encoded}"


def make_sparkline_img(values, width=90, height=24, color=None, fill_color=None, cls="spark-img"):
    svg = make_sparkline_svg(values, width, height, color, fill_color)
    return html.Img(src=svg_to_data_url(svg), className=cls, style={'display': 'block'})


# Sparklines are pure functions of (values, width, height). The same bank/metric
# sparkline is otherwise re-rendered (SVG-built + base64-encoded) on every single
# visit to a date/bank, and the All-Metrics detail builds 198 of them at once.
# Memoize the expensive bit -- the data-URL string -- keyed by a hashable tuple
# of the rounded values. lru_cache is process-wide and the inputs are static
# between data refreshes, so the cache stays warm for the dyno's life.
from functools import lru_cache as _lru_cache


@_lru_cache(maxsize=8192)
def _sparkline_data_url_cached(values_key, width, height):
    svg = make_sparkline_svg(list(values_key), width, height)
    return svg_to_data_url(svg)


def make_sparkline_img_cached(values, width=90, height=24, cls="spark-img"):
    """Memoized sparkline <img>. Values are rounded into a hashable key so tiny
    float noise does not bust the cache; identical sparklines are encoded once."""
    key = tuple(round(float(v), 6) if (v is not None and not pd.isna(v)) else None
                for v in values)
    src = _sparkline_data_url_cached(key, width, height)
    return html.Img(src=src, className=cls, style={'display': 'block'})


def make_percentile_arc_svg(pct, size=OVERVIEW_GAUGE_SIZE):
    """270-degree arc gauge for a 0-100 percentile."""
    import math
    if pct is None or pd.isna(pct):
        pct = None
    cx = cy = size / 2.0
    radius = size / 2.0 - size * 0.085
    stroke_w = size * 0.10
    start_deg, sweep_deg = 135.0, 270.0

    def _pt(deg):
        rad = math.radians(deg)
        return cx + radius * math.cos(rad), cy + radius * math.sin(rad)

    sx, sy = _pt(start_deg)
    ex_full, ey_full = _pt(start_deg + sweep_deg)
    track = (f'<path d="M {sx:.2f} {sy:.2f} A {radius:.2f} {radius:.2f} 0 1 1 '
             f'{ex_full:.2f} {ey_full:.2f}" fill="none" stroke="{CS["neutral_light"]}" '
             f'stroke-width="{stroke_w:.2f}" stroke-linecap="round"/>')
    if pct is None:
        arc = ''
        label = "\u2014"
        col = CS['neutral']
    else:
        p = max(0.0, min(100.0, float(pct)))
        col = CS['good'] if p >= 67 else (CS['warn'] if p >= 33 else CS['bad'])
        deg = sweep_deg * (p / 100.0)
        large = 1 if deg > 180 else 0
        px, py = _pt(start_deg + deg)
        arc = (f'<path d="M {sx:.2f} {sy:.2f} A {radius:.2f} {radius:.2f} 0 {large} 1 '
               f'{px:.2f} {py:.2f}" fill="none" stroke="{col}" '
               f'stroke-width="{stroke_w:.2f}" stroke-linecap="round"/>')
        label = f"{p:.0f}"
    return (f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" '
            f'xmlns="http://www.w3.org/2000/svg">{track}{arc}'
            f'<text x="{cx}" y="{cy - size * 0.02:.1f}" text-anchor="middle" dominant-baseline="central" '
            f'font-family="Inter, sans-serif" font-size="{size * 0.32:.1f}" font-weight="700" fill="{CS["text"]}">{label}</text>'
            f'<text x="{cx}" y="{cy + size * 0.22:.1f}" text-anchor="middle" dominant-baseline="central" '
            f'font-family="Inter, sans-serif" font-size="{size * 0.13:.1f}" font-weight="500" fill="{CS["text3"]}">pctl</text>'
            f'</svg>')


def make_percentile_arc_img(pct, size=OVERVIEW_GAUGE_SIZE, cls="pct-arc"):
    svg = make_percentile_arc_svg(pct, size)
    return html.Img(src=svg_to_data_url(svg), className=cls, style={'display': 'block'})


def compute_period_deltas(df, bank, metric, current_date):
    bd = df[df['Bank'] == bank].sort_values('Date')
    if bd.empty: return (None, None)
    dates = list(bd['Date'])
    try:
        idx = dates.index(pd.Timestamp(current_date))
    except ValueError:
        return (None, None)
    prev_dt = dates[idx - 1] if idx >= 1 else None
    qoq_val = (bd.iloc[idx - 1][metric]
               if idx >= 1 and 75 <= (pd.Timestamp(current_date) - prev_dt).days <= 100
               else None)
    target_yoy = pd.Timestamp(current_date) - pd.DateOffset(years=1)
    yoy_val = None
    best_j = None
    best_diff = None
    for j in range(idx):
        d = dates[j]
        diff = abs((d - target_yoy).days)
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_j = j
    if best_j is not None and best_diff <= PRIOR_PERIOD_TOLERANCE_DAYS:
        yoy_val = bd.iloc[best_j][metric]
    return (qoq_val, yoy_val)


def compute_peer_rank(df, bank, metric, date, cohort=None):
    """Direction-aware rank of `bank` among `cohort` for one metric on one
    date. Returns (rank, total, percentile) -- rank 1 = best, percentile 100 =
    best -- or (None, None, None) when unrankable. INVERSE_METRICS flip the
    ordering so lower-is-better metrics rank correctly."""
    snap = df[df['Date'] == pd.Timestamp(date)]
    if cohort is not None:
        snap = snap[snap['Bank'].isin(cohort)]
    if snap.empty or metric not in snap.columns:
        return (None, None, None)
    snap = snap.drop_duplicates(subset=['Bank'], keep='last')
    row = snap[snap['Bank'] == bank]
    if row.empty:
        return (None, None, None)
    v = row.iloc[0][metric]
    if v is None or pd.isna(v):
        return (None, None, None)
    others = [x for x in snap.loc[snap['Bank'] != bank, metric].tolist()
              if x is not None and not pd.isna(x)]
    if not others:
        return (None, None, None)
    inverse = is_inverse_metric(metric)
    better = sum(1 for x in others if ((x < v) if inverse else (x > v)))
    total = len(others) + 1
    rank = better + 1
    percentile = 100.0 * (total - rank) / (total - 1) if total > 1 else None
    return (rank, total, percentile)


def get_sparkline_series(df, bank, metric, lookback_quarters=8):
    bd = df[df['Bank'] == bank].sort_values('Date')
    if bd.empty or metric not in bd.columns:
        return []
    return bd[metric].tail(lookback_quarters).tolist()


class FDICDataUnavailableError(RuntimeError):
    pass


# =============================================================================
# FDIC API CLIENT
# =============================================================================
class FDICAPIClient:
    def __init__(self):
        self.base_url = BASE_URL

    def _get(self, ep, params, attempts=4):
        last_error = None
        for attempt in range(1, attempts + 1):
            retry_after = None
            try:
                r = requests.get(
                    f"{self.base_url}/{ep}",
                    params=params,
                    headers={"Accept": "application/json"},
                    verify=False,
                    timeout=45,
                )
                if r.status_code == 429 or 500 <= r.status_code < 600:
                    ra = r.headers.get('Retry-After')
                    if ra:
                        try:
                            retry_after = float(ra)
                        except (TypeError, ValueError):
                            retry_after = None
                    last_error = f"http {r.status_code} (rate-limited/transient)"
                else:
                    r.raise_for_status()
                    payload = r.json()
                    if isinstance(payload, dict):
                        return payload
                    last_error = f"non-dict JSON payload: {type(payload).__name__}"
            except requests.exceptions.Timeout as e:
                last_error = f"timeout: {e}"
            except requests.exceptions.RequestException as e:
                last_error = f"request error: {e}"
            except (ValueError, KeyError) as e:
                last_error = f"parse error: {e}"
            if attempt < attempts:
                backoff = retry_after if retry_after is not None else min(2 ** attempt, 12)
                backoff += random.uniform(0, 0.5)
                logger.warning(f"FDIC API {ep} attempt {attempt}/{attempts} failed "
                               f"({last_error}); retrying in {backoff:.1f}s.")
                time.sleep(backoff)
        logger.warning(f"FDIC API failed after {attempts} attempts for {ep}. Last error: {last_error}")
        return {"data": [], "_error": last_error}

    def get_institutions(self, f, fields):
        payload = self._get("institutions", {"filters": f, "fields": fields, "limit": 10})
        return payload.get('data', [])

    def get_financials(self, cert, f, fields):
        flt = f"CERT:{cert}" + (f" AND {f}" if f else "")
        payload = self._get("financials", {"filters": flt, "fields": fields,
                                           "limit": 10000, "sort_by": "REPDTE",
                                           "sort_order": "ASC"})
        return payload.get('data', [])


# =============================================================================
# RAW DATA REPOSITORY (cache discipline: only COMPLETE fetches are persisted;
# on read, a cache missing any cohort bank is rejected so a stale/partial file
# can never silently shrink the peer set.)
# =============================================================================
class BankDataRepository:
    FF = FULL_FF
    INTER_BANK_DELAY = 0.4  # seconds between per-bank FDIC calls (be polite)

    def __init__(self):
        self.api = FDICAPIClient()

    @staticmethod
    def _bank_set_hash():
        """Stable hash of (cohort certs + cache schema version) so cache files
        auto-invalidate when banks are added/removed OR the projected field set
        changes. Sorting the certs makes the hash insensitive to BANK_INFO
        ordering changes."""
        certs = ",".join(sorted(b["cert"] for b in BANK_INFO))
        return hashlib.md5(f"{certs}|{CACHE_SCHEMA_VERSION}".encode("utf-8")).hexdigest()[:10]

    def _cp(self, s, e):
        return os.path.join(CACHE_DIR, f"bank_data_{s}_{e}_{self._bank_set_hash()}.json")

    def _expected_names(self):
        return {b["display"] for b in BANK_INFO}

    def _is_complete(self, payload):
        try:
            fins = payload.get('financials_data', {})
            names = {normalize_bank_name(n) for n in fins.keys()}
            return self._expected_names().issubset(names) and all(fins.values())
        except (AttributeError, TypeError):
            return False

    def _lc(self, s, e):
        """Load cache for the exact window; if absent, fall back to the most
        recent COMPLETE cache for this cohort hash (covers redeploys earlier in
        the same quarter). Incomplete caches are rejected outright."""
        p = self._cp(s, e)
        candidates = [p] if os.path.exists(p) else []
        if not candidates:
            pattern = os.path.join(CACHE_DIR, f"bank_data_*_{self._bank_set_hash()}.json")
            candidates = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
        for path in candidates:
            try:
                with open(path) as f:
                    data = json.load(f)
            except (json.JSONDecodeError, OSError) as exc:
                logger.warning(f"Cache load failed for {path}: {exc}")
                continue
            if self._is_complete(data):
                logger.info(f"Serving COMPLETE cached FDIC dataset: {path}")
                return data
            logger.info(f"Rejecting incomplete cache {path}; will refetch.")
        return None

    def _sc(self, d, s, e):
        try:
            with open(self._cp(s, e), 'w') as f:
                json.dump(d, f)
        except OSError as exc:
            logger.warning(f"Cache save failed: {exc}")

    def fetch_data(self, bi, sd, ed, progress=None):
        def _p(msg):
            if progress:
                try:
                    progress(msg)
                except Exception:
                    pass

        c = self._lc(sd, ed)
        if c:
            _p("Loaded FDIC dataset from cache.")
            return c

        inst, fins = {}, {}
        fails = 0
        expected_count = len(bi)
        for i, b in enumerate(bi, 1):
            disp = b.get('display', b.get('cert'))
            _p(f"Fetching {disp} ({i}/{expected_count})\u2026")
            try:
                ii = self.api.get_institutions(f'CERT:{b["cert"]}', "NAME,CERT")
                if not ii:
                    logger.warning(f"No institution record for cert {b['cert']} ({disp}).")
                    fails += 1
                    continue
                bk = ii[0]
                if not (isinstance(bk, dict) and 'data' in bk):
                    fails += 1
                    continue
                bd = bk['data']
                if 'NAME' not in bd:
                    fails += 1
                    continue
                inst[bd['NAME']] = bd
                fn = self.api.get_financials(bd['CERT'], f"REPDTE:[{sd} TO {ed}]", self.FF)
                fins[bd['NAME']] = [f['data'] for f in fn if isinstance(f, dict) and 'data' in f]
                if not fins[bd['NAME']]:
                    logger.warning(f"Zero financial rows for cert {b['cert']} ({disp}).")
            except (KeyError, TypeError, ValueError) as exc:
                logger.warning(f"Fetch error for cert {b.get('cert')} ({disp}): {exc}")
                fails += 1
            if i < expected_count:
                time.sleep(self.INTER_BANK_DELAY)

        if not inst:
            raise FDICDataUnavailableError(
                f"FDIC BankFind API returned no usable data for any of the "
                f"{len(bi)} requested banks. Verify network connectivity and "
                f"FDIC API status, then retry. (Synthetic fallback has been "
                f"removed by design -- this dashboard displays real data only.)"
            )

        result = {'institutions_data': inst, 'financials_data': fins}

        if fails > 0 or len(fins) < expected_count:
            logger.warning(
                f"FDIC fetch incomplete: {expected_count - len(fins)} of {expected_count} "
                f"banks missing. Rendering available real data; intentionally not "
                f"caching. (pid={os.getpid()})")
            return result

        logger.info(f"FDIC fetch COMPLETE: all {expected_count} banks. Writing cache. (pid={os.getpid()})")
        self._sc(result, sd, ed)
        return result


# =============================================================================
# METRICS CALCULATOR -- one row per bank-quarter, every METRIC_ORDER column.
# Direct FDIC ratios pass through untouched; computed metrics follow the exact
# formulas documented in METRIC_DEFINITIONS (the definitions are the spec).
# =============================================================================
class BankMetricsCalculator:
    def __init__(self):
        # CECL adjustment coverage diagnostics (populated during calculate_metrics).
        self.cecl_window_rows = 0       # bank-quarters inside the transition window
        self.cecl_applied_rows = 0      # bank-quarters where an add-back was applied
        self.cecl_primary_samples = []  # [(REPDTE, addback$000s)] for the primary bank

    @staticmethod
    def _sf(v):
        if v is None:
            return None
        if isinstance(v, float) and pd.isna(v):
            return None
        try:
            f = float(v)
            if not np.isfinite(f):
                return None
            return f
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _z(v):
        return 0.0 if v is None else v

    @staticmethod
    def _sum_if_any(*vals):
        """Sum treating None as zero, but only when at least one component is
        present; all-None propagates None (never fabricate a zero)."""
        present = [v for v in vals if v is not None]
        if not present:
            return None
        return float(sum(present))

    def _cecl_addback(self, fin):
        """Dollar amount ($000s) of the CECL transitional add-back currently
        embedded in Tier 1 capital for this bank-quarter, to be SUBTRACTED from
        Tier 1 in the credit-concentration denominator (OCC Bull. 2020-90; Fed
        SR 20-8). Returns a non-negative float, or None when no adjustment
        applies (flag off, outside the transition window, or source fields
        absent -> caller falls back to as-reported Tier 1 + ACL).

        add-back = RC-R Pt I ln 2 retained earnings - RC item 26.a retained
        earnings (already embeds the bank's elected phase-in %). A direct
        L&L-attributable field (RC-O Memo 5 style) is preferred when configured.
        """
        if not APPLY_CECL_TRANSITION_ADJUSTMENT:
            return None
        repdte = str(fin.get('REPDTE', ''))
        if not repdte or not (CECL_TRANSITION_START <= repdte <= CECL_TRANSITION_END):
            return None
        # Preferred: a single direct add-back field, if configured/present.
        for f in CECL_DIRECT_ADDBACK_FIELDS:
            v = self._sf(fin.get(f))
            if v is not None:
                return max(0.0, v)
        # Fallback: regulatory retained earnings minus book retained earnings.
        reg = None
        for f in CECL_REG_RE_FIELDS:
            reg = self._sf(fin.get(f))
            if reg is not None:
                break
        book = None
        for f in CECL_BOOK_RE_FIELDS:
            book = self._sf(fin.get(f))
            if book is not None:
                break
        if reg is None or book is None:
            return None
        return max(0.0, reg - book)

    # ------------------------------------------------------------------ rows
    def _br(self, bank_name, fin):
        """Base row: identity, internal balance fields (underscore-prefixed,
        dropped before display), and every metric the FDIC publishes directly."""
        sf = self._sf
        return {
            'Bank': bank_name,
            'Date': fin.get('REPDTE'),
            # Internal balances ($000s) for computed metrics
            '_ta': sf(fin.get('ASSET')), '_td': sf(fin.get('DEP')),
            '_bd': sf(fin.get('BRO')), '_tl': sf(fin.get('LNLSGR')),
            '_nl': sf(fin.get('LNLSNET')), '_sc': sf(fin.get('SC')),
            '_ea': sf(fin.get('ERNAST')), '_rwa': sf(fin.get('RWAJ')),
            '_eq': sf(fin.get('EQ')), '_t1': sf(fin.get('RBCT1J')),
            '_acl': sf(fin.get('LNATRES')), '_na': sf(fin.get('NALNLS')),
            '_oreo': sf(fin.get('OREOTH')), '_p30': sf(fin.get('P3LNLS')),
            '_p90': sf(fin.get('P9LNLS')), '_ni': sf(fin.get('NETINC')),
            '_niq': sf(fin.get('NETINCQ')), '_gco_ytd': sf(fin.get('DRLNLS')),
            '_gcoq': sf(fin.get('DRLNLSQ')), '_rec_ytd': sf(fin.get('CRLNLS')),
            '_recq': sf(fin.get('CRLNLSQ')), '_ncoq': sf(fin.get('NTLNLSQ')),
            # Concentration exposures ($000s)
            '_re': sf(fin.get('LNRE')), '_cons': sf(fin.get('LNRECONS')),
            '_c14': sf(fin.get('LNRECNFM')), '_cot': sf(fin.get('LNRECNOT')),
            '_farm': sf(fin.get('LNREAG')), '_res14': sf(fin.get('LNRERES')),
            '_heloc': sf(fin.get('LNRELOC')), '_cl1': sf(fin.get('LNRERSFM')),
            '_cljr': sf(fin.get('LNRERSF2')), '_mf': sf(fin.get('LNREMULT')),
            '_nfnr': sf(fin.get('LNRENRES')), '_nfnrow': sf(fin.get('LNRENROW')),
            '_nfnrnoo': sf(fin.get('LNRENROT')), '_comre': sf(fin.get('LNCOMRE')),
            '_ci': sf(fin.get('LNCI')), '_agl': sf(fin.get('LNAG')),
            '_con': sf(fin.get('LNCON')), '_cards': sf(fin.get('LNCRCD')),
            '_auto': sf(fin.get('LNAUTO')), '_oth': sf(fin.get('LNOTHER')),
            # FDIC-provided ratios (displayed as-is)
            'Net Charge-Offs / Total Loans & Leases': sf(fin.get('NTLNLSR')),
            'ACL / Total Loans & Leases': sf(fin.get('LNATRESR')),
            'Earnings Coverage of Net Loan Charge-Offs': sf(fin.get('IDERNCVR')),
            'Loan and Lease Loss Provision to Net Charge-Offs': sf(fin.get('ELNANTR')),
            'Net Loans and Leases to Deposits': sf(fin.get('LNLSDEPR')),
            'Net Loans and Leases to Assets': sf(fin.get('LNLSNTV')),
            'Return on Assets': sf(fin.get('ROA')),
            'Quarterly Return on Assets': sf(fin.get('ROAQ')),
            'Return on Equity': sf(fin.get('ROE')),
            'Quarterly Return on Equity': sf(fin.get('ROEQ')),
            'Leverage (Core Capital) Ratio': sf(fin.get('RBC1AAJ')),
            'Total Risk-Based Capital Ratio': sf(fin.get('RBCRWAJ')),
            'Efficiency Ratio': sf(fin.get('EEFFR')),
            'Earning Assets / Total Assets': sf(fin.get('ERNASTR')),
            'Net Interest Margin': sf(fin.get('NIMY')),
            'Common Equity Tier 1 (CET1) Ratio': sf(fin.get('IDT1CER')),
            'Tier 1 Risk-Based Capital Ratio': sf(fin.get('IDT1RWAJR')),
            'Cost of Funding Earning Assets (YTD)': sf(fin.get('INTEXPY')),
            'Cost of Funding Earning Assets (Quarterly)': sf(fin.get('INTEXPYQ')),
            'Noninterest Income to Average Assets': sf(fin.get('NONIIR')),
            'Pretax Return on Assets': sf(fin.get('ROAPTX')),
            'Noninterest Expense to Average Assets': sf(fin.get('NONIXR')),
            'Loan Loss Reserve / Noncurrent Loans': sf(fin.get('LNRESNCR')),
            'Volatile Liabilities to Total Assets': sf(fin.get('VOLIABR')),
            'Net Operating Income to Assets': sf(fin.get('NOIJY')),
            'Salaries and Benefits to Average Assets': sf(fin.get('ESALR')),
            'Interest Income to Average Assets': sf(fin.get('INTINCR')),
            'Interest Expense to Average Assets': sf(fin.get('EINTEXPR')),
            'Provision for Credit Losses to Average Assets': sf(fin.get('ELNATRR')),
            'Yield on Earning Assets': sf(fin.get('INTINCY')),
        }

    def _kf(self, r, fin):
        r['Total Assets'] = r['_ta']
        r['Total Deposits'] = r['_td']
        r['Gross Loans & Leases'] = r['_tl']
        # Net Loans & Leases -- LNLSNET, the numerator FDIC uses for the
        # LNLSDEPR (Net Loans / Deposits) ratio. Already pulled into _nl in _br().
        r['Net Loans & Leases'] = r['_nl']
        r['Total Securities'] = r['_sc']
        r['Total Earning Assets'] = r['_ea']
        r['Total Equity Capital'] = r['_eq']
        r['Tier 1 Capital'] = r['_t1']
        r['Risk-Weighted Assets'] = r['_rwa']
        r['Net Income (YTD)'] = r['_ni']
        r['Net Income (Quarter)'] = r['_niq']
        r['Allowance for Credit Losses'] = r['_acl']
        r['Gross Charge-Offs (YTD)'] = r['_gco_ytd']
        r['Gross Charge-Offs (Quarter)'] = r['_gcoq']
        r['Gross Recoveries (YTD)'] = r['_rec_ytd']
        r['Gross Recoveries (Quarter)'] = r['_recq']
        # Noncurrent loans is a sum, so missing components are treated as zero
        # only when at least one component is present.
        if r['_na'] is not None or r['_p90'] is not None:
            r['Noncurrent Loans'] = self._z(r['_na']) + self._z(r['_p90'])
        else:
            r['Noncurrent Loans'] = None
        r['Brokered Deposits'] = r['_bd']

    def _cb(self, r, fin):
        """Concentration denominator: Tier 1 + ACL, with Tier 1 reduced by the
        CECL transitional add-back during the 2020-2024 window (OCC Bulletin
        2020-90; Fed SR 20-8) so the reserve is not double-counted -- matching
        the UBPR Page 7B methodology. Fields absent -> safe no-op (raw T1+ACL).
        Returns the denominator ($000s) or None when it cannot be formed."""
        t1, acl = r['_t1'], r['_acl']
        if t1 is None or acl is None:
            return None
        repdte = str(fin.get('REPDTE', ''))
        in_window = bool(repdte) and (CECL_TRANSITION_START <= repdte <= CECL_TRANSITION_END)
        if in_window:
            self.cecl_window_rows += 1
        addback = self._cecl_addback(fin)
        adj_t1 = t1
        if addback is not None and addback > 0:
            candidate = t1 - addback
            if candidate > 0:
                adj_t1 = candidate
                self.cecl_applied_rows += 1
                if r.get('Bank') == PRIMARY_BANK_DISPLAY_NAME and len(self.cecl_primary_samples) < 8:
                    self.cecl_primary_samples.append((repdte, addback))
            # else: add-back exceeds Tier 1 (implausible) -> fall back unadjusted.
        base = adj_t1 + acl
        return base if base > 0 else None

    def _cc(self, r, fin):
        cb = self._cb(r, fin)
        z, sa = self._z, self._sum_if_any
        r['Real Estate Loans to Tier 1 + ACL'] = safe_div(r['_re'], cb)
        r['RE Construction and Land Development to Tier 1 + ACL'] = safe_div(r['_cons'], cb)
        r['1-4 Family Construction to Tier 1 + ACL'] = safe_div(r['_c14'], cb)
        r['Other Construction & Land Dev to Tier 1 + ACL'] = safe_div(r['_cot'], cb)
        r['Secured by Farmland to Tier 1 + ACL'] = safe_div(r['_farm'], cb)
        r['1-4 Family Residential to Tier 1 + ACL'] = safe_div(r['_res14'], cb)
        r['Revolving Home Equity to Tier 1 + ACL'] = safe_div(r['_heloc'], cb)
        r['Closed-End 1st Lien to Tier 1 + ACL'] = safe_div(r['_cl1'], cb)
        r['Closed-End Jr Lien to Tier 1 + ACL'] = safe_div(r['_cljr'], cb)
        r['Multifamily RE to Tier 1 + ACL'] = safe_div(r['_mf'], cb)
        r['Non-Farm Non-Residential RE to Tier 1 + ACL'] = safe_div(r['_nfnr'], cb)
        r['NFNR: Owner Occupied to Tier 1 + ACL'] = safe_div(r['_nfnrow'], cb)
        r['NFNR: Non-Owner Occupied to Tier 1 + ACL'] = safe_div(r['_nfnrnoo'], cb)
        # UBPR Total CRE = Construction + Multifamily + NFNR total + LNCOMRE.
        cre = sa(r['_cons'], r['_mf'], r['_nfnr'], r['_comre'])
        r['Commercial RE to Tier 1 + ACL'] = safe_div(cre, cb)
        # Interagency NOO CRE = Construction + Multifamily + NFNR NOO + LNCOMRE.
        noo = sa(r['_cons'], r['_mf'], r['_nfnrnoo'], r['_comre'])
        r['_noo_cre'] = noo  # dollar exposure retained for the 3-yr growth pass
        r['Non-Owner Occupied CRE to Tier 1 + ACL'] = safe_div(noo, cb)
        r['C&I Loans to Tier 1 + ACL'] = safe_div(r['_ci'], cb)
        r['Loans to Individuals to Tier 1 + ACL'] = safe_div(r['_con'], cb)
        r['Credit Cards to Tier 1 + ACL'] = safe_div(r['_cards'], cb)
        r['Auto Loans to Tier 1 + ACL'] = safe_div(r['_auto'], cb)
        r['Agriculture Loans to Tier 1 + ACL'] = safe_div(r['_agl'], cb)
        r['Loans to NDFIs and Other to Tier 1 + ACL'] = safe_div(r['_oth'], cb)

    def _aq(self, r, fin):
        tl = r['_tl']
        r['30-89 DPD / Total Loans'] = safe_div(r['_p30'], tl)
        r['90+ DPD / Total Loans'] = safe_div(r['_p90'], tl)
        r['Nonaccrual / Total Loans'] = safe_div(r['_na'], tl)
        nc = self._sum_if_any(r['_na'], r['_p90'])
        r['90+ DPD & Nonaccrual / Total Loans'] = safe_div(nc, tl)
        # Broadest NPA measure: (nonaccrual + OREO + 90+ DPD) / (loans + OREO).
        npa = self._sum_if_any(r['_na'], r['_oreo'], r['_p90'])
        denom = self._sum_if_any(tl, r['_oreo']) if tl is not None else None
        r['Nonaccrual & OREO / Total Loans & OREO'] = safe_div(npa, denom)
        # Coverage MULTIPLES (x, not %).
        r['ACL / Nonaccrual Loans'] = safe_div(r['_acl'], r['_na'], scale=1.0)
        r['ACL / 90+ DPD & Nonaccrual'] = safe_div(r['_acl'], nc, scale=1.0)

    def _aq_segments(self, r, fin):
        """Segment-level stock rates, dollars, and annualized NCO flow rates --
        generated straight from ASSET_QUALITY_SEGMENTS so names, fields, and
        math can never drift apart (see the table's docstring)."""
        sf = self._sf
        repdte = str(fin.get('REPDTE', ''))
        try:
            month = int(repdte[4:6])
            quarter = max(1, min(4, (month + 2) // 3))
        except (ValueError, IndexError):
            quarter = 4
        for (label, bal_f, na_f, p9_f, p3_f, nt_f, cavg5_f) in ASSET_QUALITY_SEGMENTS:
            bal = sf(fin.get(bal_f))
            na = sf(fin.get(na_f))
            p9 = sf(fin.get(p9_f))
            p3 = sf(fin.get(p3_f))
            nt = sf(fin.get(nt_f))
            r[_nm_na_rate(label)] = safe_div(na, bal)
            r[_nm_p9_rate(label)] = safe_div(p9, bal)
            r[_nm_p3_rate(label)] = safe_div(p3, bal)
            r[_nm_na_usd(label)] = na
            r[_nm_p9_usd(label)] = p9
            r[_nm_p3_usd(label)] = p3
            r[_nm_nco_usd(label)] = nt
            if cavg5_f is not None:
                cavg5 = sf(fin.get(cavg5_f))
                # UBPR PCTOFANN: YTD NCO annualized over the 5-quarter average.
                ann_nco = nt * (4.0 / quarter) if nt is not None else None
                r[_nm_nco_rate(label)] = safe_div(ann_nco, cavg5)

    def _gr(self, r, fin):
        r[GROSS_RECOVERY_RATE_METRIC] = safe_div(r['_rec_ytd'], r['_gco_ytd'])

    def _bk(self, r, fin):
        sf = self._sf
        r['Core Deposits to Total Deposits'] = safe_div(sf(fin.get('COREDEP')), r['_td'])
        r['Noninterest-Bearing Deposits to Total Deposits'] = safe_div(sf(fin.get('DEPNIDOM')), r['_td'])
        r['Brokered Deposits to Total Deposits'] = safe_div(r['_bd'], r['_td'])

    def _dp(self, r, fin):
        """PPNR / Avg Assets from the four published avg-asset ratios (UBPR Pg1
        #6 arithmetic): II - IE + NonII - NonIE. Any missing component -> None."""
        comps = [r.get('Interest Income to Average Assets'),
                 r.get('Interest Expense to Average Assets'),
                 r.get('Noninterest Income to Average Assets'),
                 r.get('Noninterest Expense to Average Assets')]
        if any(c is None or pd.isna(c) for c in comps):
            r['Pre-Provision Net Revenue to Average Assets'] = None
        else:
            ii, ie, nii, nie = comps
            r['Pre-Provision Net Revenue to Average Assets'] = ii - ie + nii - nie

    # ------------------------------------------------- time-series second pass
    @staticmethod
    def _nearest_prior(dates, idx, years_back):
        """Index of the observation closest to (dates[idx] - years_back), or
        None if nothing lands within the prior-period tolerance."""
        target = dates[idx] - pd.DateOffset(years=years_back)
        best_j, best_diff = None, None
        for j in range(idx):
            diff = abs((dates[j] - target).days)
            if best_diff is None or diff < best_diff:
                best_diff, best_j = diff, j
        if best_j is not None and best_diff <= PRIOR_PERIOD_TOLERANCE_DAYS:
            return best_j
        return None

    def _cg(self, bdf):
        """YoY growth rates (and the 3-yr NOO CRE growth) for one bank's
        date-sorted frame. Dollar-balance growth: NOT CECL-affected."""
        dates = list(bdf['Date'])
        specs = [('_ta', 'Total Asset Growth Rate', 1),
                 ('_t1', 'Tier 1 Capital Growth Rate', 1),
                 ('_nl', 'Net Loan Growth Rate', 1),
                 ('_noo_cre', 'Non-Owner Occupied CRE 3-Year Growth Rate', 3)]
        for src, name, yrs in specs:
            out = [None] * len(bdf)
            if src in bdf.columns:
                vals = list(bdf[src])
                for i in range(len(bdf)):
                    j = self._nearest_prior(dates, i, yrs)
                    if j is None:
                        continue
                    cur, prior = self._sf(vals[i]), self._sf(vals[j])
                    if cur is None or prior is None or prior <= 0:
                        continue
                    out[i] = ((cur / prior) - 1.0) * 100.0
            bdf[name] = out
        return bdf

    def _pp(self, bdf):
        """Rolling 4-quarter NCOs as % of current ACL. Requires four CONTIGUOUS
        quarters (75-100 day spacing); otherwise N/A -- never a partial sum."""
        dates = list(bdf['Date'])
        ncoq = list(bdf['_ncoq']) if '_ncoq' in bdf.columns else [None] * len(bdf)
        acl = list(bdf['_acl']) if '_acl' in bdf.columns else [None] * len(bdf)
        out = [None] * len(bdf)
        for i in range(len(bdf)):
            if i < 3:
                continue
            window = [self._sf(ncoq[k]) for k in (i - 3, i - 2, i - 1, i)]
            if any(v is None for v in window):
                continue
            contiguous = all(75 <= (dates[k] - dates[k - 1]).days <= 100
                             for k in range(i - 2, i + 1))
            if not contiguous:
                continue
            a = self._sf(acl[i])
            out[i] = safe_div(sum(window), a)
        bdf['Net Charge-Offs / ACL'] = out
        return bdf

    # ----------------------------------------------------------------- driver
    def calculate_metrics(self, financials_data):
        frames = []
        for raw_name, fins in financials_data.items():
            bank = normalize_bank_name(raw_name)
            rows = []
            for fin in sorted(fins, key=lambda f: str(f.get('REPDTE', ''))):
                if not isinstance(fin, dict) or not fin.get('REPDTE'):
                    continue
                r = self._br(bank, fin)
                self._kf(r, fin)
                self._cc(r, fin)
                self._aq(r, fin)
                self._aq_segments(r, fin)
                self._gr(r, fin)
                self._bk(r, fin)
                self._dp(r, fin)
                rows.append(r)
            if not rows:
                continue
            bdf = pd.DataFrame(rows)
            bdf['Date'] = pd.to_datetime(bdf['Date'], format='%Y%m%d', errors='coerce')
            bdf = bdf.dropna(subset=['Date']).sort_values('Date')
            # FDIC normally returns one row per REPDTE; dedupe defensively.
            bdf = bdf.drop_duplicates(subset=['Bank', 'Date'], keep='last').reset_index(drop=True)
            bdf = self._cg(bdf)
            bdf = self._pp(bdf)
            frames.append(bdf)
        if not frames:
            return pd.DataFrame()
        df = pd.concat(frames, ignore_index=True)
        return df.sort_values(['Bank', 'Date']).reset_index(drop=True)

    def cecl_status(self):
        cov = (100.0 * self.cecl_applied_rows / self.cecl_window_rows
               if self.cecl_window_rows else None)
        return {'window_rows': self.cecl_window_rows,
                'applied_rows': self.cecl_applied_rows,
                'coverage_pct': cov,
                'primary_samples': list(self.cecl_primary_samples)}


# =============================================================================
# SERVICE LAYER
# =============================================================================
class BankDataService:
    def __init__(self):
        self.repo = BankDataRepository()
        self.calc = BankMetricsCalculator()

    def get_metrics_data(self, sd=DEFAULT_START_DATE, ed=DEFAULT_END_DATE, progress=None):
        raw = self.repo.fetch_data(BANK_INFO, sd, ed, progress=progress)
        if progress:
            try:
                progress(f"Computing {len(METRIC_ORDER)} UBPR-aligned metrics\u2026")
            except Exception:
                pass
        df = self.calc.calculate_metrics(raw.get('financials_data', {}))
        if df.empty:
            return df
        # Internal underscore columns never leave the service layer.
        drop = [c for c in df.columns if c.startswith('_')]
        return df.drop(columns=drop)


# =============================================================================
# EXCEL EXPORT -- every period, every metric, one sheet per category, for any
# bank in the cohort (defaults to the primary bank).
# =============================================================================
def _safe_sheet_title(name, max_len=31):
    """Excel sheet titles: <=31 chars, no []:*?/\\ characters."""
    bad = set('[]:*?/\\')
    clean = ''.join(('-' if ch in bad else ch) for ch in str(name)).strip()
    return (clean[:max_len]) if clean else "Sheet"


def build_bank_export(df, bank_display=PRIMARY_BANK_DISPLAY_NAME):
    """All-periods Excel workbook for one bank: a sheet per metric category,
    metrics as rows, reporting dates as columns (oldest -> newest), values
    formatted exactly as the dashboard shows them. Returns workbook bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    bdf = df[df['Bank'] == bank_display].sort_values('Date')
    dates = list(bdf['Date'])

    wb = Workbook()
    wb.remove(wb.active)

    hdr_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='005EB8', end_color='005EB8', fill_type='solid')
    name_font = Font(name='Calibri', size=10, bold=True)
    val_font = Font(name='Calibri', size=10)
    thin = Side(style='thin', color='D9E2EC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cat_name, cat_metrics in METRIC_CATEGORIES:
        present = [m for m in cat_metrics if m in bdf.columns]
        if not present:
            continue
        ws = wb.create_sheet(_safe_sheet_title(cat_name))
        ws.cell(row=1, column=1, value='Metric')
        for j, d in enumerate(dates, start=2):
            c = ws.cell(row=1, column=j, value=d.strftime('%m/%d/%Y'))
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')
            c.border = border
        c0 = ws.cell(row=1, column=1)
        c0.font = hdr_font
        c0.fill = hdr_fill
        c0.border = border
        ws.freeze_panes = 'B2'
        ws.column_dimensions['A'].width = 46
        for i, m in enumerate(present, start=2):
            nc = ws.cell(row=i, column=1, value=m)
            nc.font = name_font
            nc.border = border
            series = list(bdf[m])
            for j, v in enumerate(series, start=2):
                vc = ws.cell(row=i, column=j, value=fmt_val(v, m, with_unit=True))
                vc.font = val_font
                vc.alignment = Alignment(horizontal='right')
                vc.border = border
        for j in range(2, len(dates) + 2):
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# PAGE TEMPLATE + DESIGN SYSTEM (single source of CSS; Dash tokens preserved)
# =============================================================================
INDEX_STRING = """<!DOCTYPE html>
<html>
<head>
    {%metas%}
    <title>{%title%}</title>
    {%favicon%}
    {%css%}
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        :root{
            --jpm:#005EB8; --jpm-dark:#003B73; --jpm-light:#2F7FD3;
            --ink:#0f172a; --ink2:#475569; --ink3:#64748b;
            --bg:#f4f6f9; --card:#ffffff;
            --line:rgba(15,23,42,0.06); --line2:rgba(15,23,42,0.12);
            --good:#16a34a; --bad:#ef4444; --warn:#f59e0b; --neutral:#64748b;
            --r-lg:14px; --r-md:10px; --r-sm:7px;
            --sh-1:0 1px 2px rgba(15,23,42,.05),0 1px 1px rgba(15,23,42,.03);
            --sh-2:0 8px 24px -8px rgba(2,32,71,.16),0 2px 6px rgba(2,32,71,.06);
        }
        *,*::before,*::after{box-sizing:border-box}
        html{-webkit-text-size-adjust:100%}
        body{margin:0;background:var(--bg);color:var(--ink);
            font-family:'Inter',-apple-system,'Segoe UI',Roboto,sans-serif;
            font-size:14px;line-height:1.45;
            -webkit-font-smoothing:antialiased;-moz-osx-font-smoothing:grayscale}
        ::selection{background:rgba(0,94,184,.18)}
        ::-webkit-scrollbar{width:10px;height:10px}
        ::-webkit-scrollbar-track{background:transparent}
        ::-webkit-scrollbar-thumb{background:#c7d2de;border-radius:6px;border:2px solid var(--bg)}
        ::-webkit-scrollbar-thumb:hover{background:#9fb2c4}
        :focus-visible{outline:2px solid var(--jpm);outline-offset:2px;border-radius:4px}

        /* ---------- header ---------- */
        .hdr{position:sticky;top:0;z-index:60;
            background:linear-gradient(135deg,#003B73 0%,#005EB8 62%,#0B4F8A 100%);
            color:#fff;box-shadow:0 6px 22px -8px rgba(2,32,71,.55);
            backdrop-filter:saturate(140%) blur(6px)}
        .hdr-inner{max-width:1480px;margin:0 auto;padding:16px 28px 14px}
        .hdr-title{margin:0;font-size:21px;font-weight:800;letter-spacing:-.015em;line-height:1.2}
        .hdr-meta{display:flex;align-items:center;flex-wrap:wrap;gap:7px 14px;
            margin-top:6px;font-size:12px;font-weight:500;color:rgba(255,255,255,.82)}
        .hdr-live{display:inline-flex;align-items:center;gap:6px;
            padding:2.5px 10px 2.5px 8px;border-radius:999px;
            background:rgba(255,255,255,.13);border:1px solid rgba(255,255,255,.22);
            font-size:11px;font-weight:600;letter-spacing:.04em;text-transform:uppercase}
        .live-dot{width:7px;height:7px;border-radius:50%;background:#4ade80;
            box-shadow:0 0 0 0 rgba(74,222,128,.65);animation:pulse 2.2s infinite}
        @keyframes pulse{0%{box-shadow:0 0 0 0 rgba(74,222,128,.65)}
            70%{box-shadow:0 0 0 7px rgba(74,222,128,0)}100%{box-shadow:0 0 0 0 rgba(74,222,128,0)}}
        .hdr-disc{margin-top:4px;font-size:11px;color:rgba(255,255,255,.6)}

        /* ---------- shell ---------- */
        .main-wrap{max-width:1480px;margin:0 auto;padding:22px 28px 56px}
        .ftr{max-width:1480px;margin:0 auto;padding:18px 28px 30px;
            font-size:11.5px;color:var(--ink3);border-top:1px solid var(--line)}

        /* ---------- cards ---------- */
        .card{background:var(--card);border:1px solid var(--line);border-radius:var(--r-lg);
            box-shadow:var(--sh-1);padding:18px 20px;
            transition:box-shadow .18s ease,transform .18s ease}
        .ch{display:flex;align-items:baseline;justify-content:space-between;
            flex-wrap:wrap;gap:6px 12px;margin-bottom:12px}
        .ct{margin:0;font-size:14.5px;font-weight:700;letter-spacing:-.01em;color:var(--ink)}
        .csub,.rng{font-size:11.5px;font-weight:500;color:var(--ink3)}

        .sec{margin-top:26px}
        .sec-head{margin:0 2px 12px}
        .sec-title{margin:0;font-size:16.5px;font-weight:800;letter-spacing:-.015em}
        .sec-sub{margin:3px 0 0;font-size:12px;color:var(--ink3)}

        /* ---------- peer controls ---------- */
        .peer-card{margin-top:20px}
        .peer-row{display:flex;align-items:flex-start;gap:14px;flex-wrap:wrap}
        .peer-dd-wrap{flex:1 1 520px;min-width:300px}
        .peer-actions{display:flex;gap:8px;padding-top:2px}
        .btn-mini{appearance:none;border:1px solid var(--line2);background:#fff;
            color:var(--ink2);font:inherit;font-size:12px;font-weight:600;
            padding:7px 14px;border-radius:var(--r-sm);cursor:pointer;
            transition:all .15s ease;white-space:nowrap}
        .btn-mini:hover{border-color:var(--jpm);color:var(--jpm);background:#f5f9fe}
        .peer-count{font-size:11.5px;color:var(--ink3);margin-top:8px}

        /* ---------- executive banner ---------- */
        .exec-grid{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));
            gap:12px;margin-top:16px}
        .exec-card{position:relative;background:var(--card);border:1px solid var(--line);
            border-radius:var(--r-md);box-shadow:var(--sh-1);padding:13px 14px 12px;
            overflow:hidden;transition:box-shadow .18s ease,transform .18s ease}
        .exec-card::before{content:"";position:absolute;left:0;top:0;bottom:0;width:3px;
            background:var(--jpm);opacity:0;transition:opacity .18s ease}
        .exec-card:hover{transform:translateY(-2px);box-shadow:var(--sh-2)}
        .exec-card:hover::before{opacity:1}
        .exec-kpi-cat{font-size:10px;font-weight:700;letter-spacing:.08em;
            text-transform:uppercase;color:var(--jpm)}
        .exec-kpi-label{font-size:11.5px;font-weight:600;color:var(--ink2);
            margin-top:2px;min-height:30px;line-height:1.3}
        .exec-kpi-val{font-size:21px;font-weight:800;letter-spacing:-.02em;
            font-variant-numeric:tabular-nums;margin-top:4px}
        .exec-kpi-row{display:flex;align-items:center;justify-content:space-between;
            gap:8px;margin-top:7px}
        .exec-rank{font-size:10.5px;font-weight:600;color:var(--ink3)}
        .spark-img{width:90px;height:24px;margin-top:8px}

        .delta-chip{display:inline-flex;align-items:center;gap:3px;
            font-size:10.5px;font-weight:700;font-variant-numeric:tabular-nums;
            padding:2px 7px;border-radius:999px;white-space:nowrap}
        .delta-chip.up{color:var(--good);background:color-mix(in srgb,var(--good) 12%,white)}
        .delta-chip.down{color:var(--bad);background:color-mix(in srgb,var(--bad) 11%,white)}
        .delta-chip.flat{color:var(--neutral);background:#f1f5f9}

        /* ---------- analysis sections ---------- */
        .control-row{display:flex;align-items:flex-end;gap:14px;flex-wrap:wrap;margin-bottom:14px}
        .ctl{display:flex;flex-direction:column;gap:5px}
        .ctl-label{font-size:10.5px;font-weight:700;letter-spacing:.07em;
            text-transform:uppercase;color:var(--ink3)}
        .peer-metric-wrap{flex:0 0 500px;max-width:610px;min-width:360px;
            transition:flex-basis .2s ease}
        .peer-def{font-size:11.5px;color:var(--ink3);line-height:1.5;
            margin:2px 2px 14px;max-width:1100px}
        .range-label{font-size:11.5px;font-weight:600;color:var(--ink3)}

        .paired-row{display:flex;gap:16px;align-items:stretch}
        .chart-col{flex:1 1 62%;min-width:0;display:flex}
        .insight-col{flex:0 0 36%;min-width:300px;display:flex}
        .chart-col .card,.insight-col .card{flex:1;min-height:432px;display:flex;flex-direction:column}
        .insight-shell{flex:1;overflow-y:auto;min-height:0}

        /* ---------- overview / analysis panels ---------- */
        .ov-top{display:flex;align-items:center;justify-content:space-between;gap:14px}
        .ov-val{font-size:30px;font-weight:800;letter-spacing:-.025em;
            font-variant-numeric:tabular-nums;line-height:1.05}
        .ov-unit{font-size:13px;font-weight:600;color:var(--ink3);margin-left:3px}
        .ov-rank{font-size:12px;font-weight:600;color:var(--ink2);margin-top:5px}
        .ov-stats{display:grid;grid-template-columns:1fr 1fr;gap:8px 14px;margin-top:14px}
        .ov-stat{padding:8px 10px;background:#f8fafc;border:1px solid var(--line);
            border-radius:var(--r-sm)}
        .ov-stat-label{font-size:10px;font-weight:700;letter-spacing:.06em;
            text-transform:uppercase;color:var(--ink3)}
        .ov-stat-val{font-size:13.5px;font-weight:700;font-variant-numeric:tabular-nums;margin-top:2px}
        .ov-mom-label{font-size:10.5px;font-weight:700;letter-spacing:.07em;
            text-transform:uppercase;color:var(--ink3);margin:16px 0 6px}
        .pct-arc{flex:0 0 auto}

        .jpm-corr-card .corr-val{font-size:30px;font-weight:800;
            font-variant-numeric:tabular-nums;letter-spacing:-.02em}
        .corr-label{font-size:11px;font-weight:700;letter-spacing:.06em;
            text-transform:uppercase;color:var(--ink3)}
        .corr-interp{font-size:12.5px;color:var(--ink2);line-height:1.55;margin-top:10px}

        .emp{color:var(--ink3);font-size:13px;padding:18px 4px}
        .warn-banner{margin-top:16px;padding:11px 16px;border-radius:var(--r-md);
            background:#fffbeb;border:1px solid #fde68a;color:#92400e;
            font-size:12.5px;font-weight:500}

        /* ---------- All-Metrics detail ---------- */
        .det-controls{display:flex;align-items:flex-end;gap:14px;flex-wrap:wrap}
        .det-bank-wrap{flex:0 0 300px;min-width:240px;transition:flex-basis .2s ease}
        .det-date-wrap{flex:0 0 190px}
        .btn-export{appearance:none;border:none;cursor:pointer;font:inherit;
            font-size:12.5px;font-weight:700;color:#fff;background:var(--jpm);
            padding:9px 18px;border-radius:var(--r-sm);
            box-shadow:0 2px 8px -2px rgba(0,94,184,.55);transition:all .15s ease}
        .btn-export:hover{background:var(--jpm-dark);transform:translateY(-1px)}
        .det-legend{font-size:11px;color:var(--ink3);margin-left:auto;align-self:center}
        .det-cat{margin-top:18px;border:1px solid var(--line);border-radius:var(--r-md);overflow:hidden}
        .det-cat-head{display:flex;align-items:center;gap:10px;padding:9px 14px;
            font-size:12px;font-weight:800;letter-spacing:.03em}
        .det-cat-count{margin-left:auto;font-size:10.5px;font-weight:600;color:var(--ink3)}
        .det-row{display:grid;grid-template-columns:minmax(230px,2.1fr) minmax(95px,.9fr) 100px minmax(86px,.7fr) minmax(86px,.7fr);
            gap:12px;align-items:center;padding:7px 14px;border-top:1px solid var(--line);
            background:#fff;font-size:12.5px}
        .det-row:nth-child(even){background:#fbfcfe}
        .det-row:hover{background:#f4f8fd}
        .det-hdr-row{font-size:10px;font-weight:700;letter-spacing:.07em;
            text-transform:uppercase;color:var(--ink3);background:#fff}
        .det-name{font-weight:600;color:var(--ink)}
        .det-val{font-weight:700;font-variant-numeric:tabular-nums;text-align:right}
        .det-delta{font-size:11.5px;font-weight:700;font-variant-numeric:tabular-nums;text-align:right}

        /* ---------- reference guide (collapsible) ---------- */
        .ref-wrap{column-width:300px;column-gap:14px;margin-top:6px}
        details.ref-cat{break-inside:avoid;background:#fff;border:1px solid var(--line);
            border-radius:var(--r-md);margin:0 0 14px;overflow:hidden}
        .ref-summary{display:flex;align-items:center;gap:10px;list-style:none;
            cursor:pointer;padding:10px 14px;user-select:none}
        .ref-summary::-webkit-details-marker{display:none}
        .ref-accent{width:4px;height:18px;border-radius:3px;flex:0 0 auto}
        .ref-cat-label{font-size:12.5px;font-weight:700}
        .ref-cat-count{margin-left:auto;font-size:10.5px;font-weight:600;color:var(--ink3)}
        .ref-chev{flex:0 0 auto;color:var(--ink3);font-size:11px;
            transition:transform .18s ease}
        details[open] .ref-chev{transform:rotate(90deg)}
        .ref-body{padding:2px 14px 12px;border-top:1px solid var(--line)}
        .ref-row{padding:8px 0;border-bottom:1px dashed var(--line)}
        .ref-row:last-child{border-bottom:none}
        .ref-name{font-size:12px;font-weight:700}
        .ref-desc{font-size:11.5px;color:var(--ink2);line-height:1.5;margin-top:2px}

        /* ---------- dropdown theming (both react-select generations) ---------- */
        .dd .Select-control,.dd [class*="-control"]{
            border:1px solid var(--line2)!important;border-radius:var(--r-sm)!important;
            min-height:38px!important;box-shadow:none!important;
            font-size:13px;transition:border-color .15s ease}
        .dd .Select-control:hover,.dd [class*="-control"]:hover{border-color:var(--jpm)!important}
        .dd .is-focused .Select-control,.dd [class*="-control"][class*="-is-focused"],
        .dd [class*="-control"]:focus-within{
            border-color:var(--jpm)!important;box-shadow:0 0 0 3px rgba(0,94,184,.14)!important}
        .dd .Select-menu-outer,.dd [class*="-menu"]{
            border:1px solid var(--line2)!important;border-radius:var(--r-sm)!important;
            box-shadow:var(--sh-2)!important;font-size:13px;z-index:80!important}
        .dd .VirtualizedSelectFocusedOption,.dd [class*="-option"][class*="-is-focused"]{
            background:#eef5fd!important;color:var(--ink)!important}
        .dd [class*="-option"][class*="-is-selected"]{
            background:var(--jpm)!important;color:#fff!important}
        /* multi-select peer chips: compact, scrollable when 18 are selected */
        .dd .Select--multi .Select-multi-value-wrapper,
        .dd [class*="-control"] [class*="-ValueContainer"]{
            max-height:88px;overflow-y:auto}
        .dd .Select--multi .Select-value,.dd [class*="-multiValue"]{
            background:#e8f2fc!important;border:1px solid rgba(0,94,184,.25)!important;
            border-radius:5px!important;color:var(--jpm-dark)!important;
            font-size:11.5px;font-weight:600}
        .dd .Select--multi .Select-value-icon:hover,
        .dd [class*="-multiValue"] [role="button"]:hover{
            background:rgba(0,94,184,.18)!important;color:var(--jpm-dark)!important}

        /* ---------- boot / error screens ---------- */
        .boot-screen{min-height:100vh;display:flex;align-items:center;justify-content:center;
            padding:28px;background:
            radial-gradient(1100px 520px at 18% -10%,rgba(0,94,184,.16),transparent 60%),
            radial-gradient(900px 480px at 100% 110%,rgba(11,79,138,.13),transparent 55%),
            var(--bg)}
        .boot-card{width:min(560px,94vw);background:#fff;border:1px solid var(--line);
            border-radius:18px;box-shadow:var(--sh-2);padding:38px 40px;text-align:center}
        .boot-mark{width:54px;height:54px;margin:0 auto;border-radius:14px;
            display:flex;align-items:center;justify-content:center;
            background:linear-gradient(135deg,#003B73,#005EB8);color:#fff;
            font-size:17px;font-weight:800;letter-spacing:.02em;
            box-shadow:0 10px 24px -8px rgba(0,94,184,.6)}
        .boot-title{margin:18px 0 4px;font-size:17px;font-weight:800;letter-spacing:-.01em}
        .boot-dots{display:flex;gap:7px;justify-content:center;margin:18px 0 14px}
        .boot-dot{width:9px;height:9px;border-radius:50%;background:var(--jpm);
            animation:bdot 1.25s ease-in-out infinite}
        .boot-dot:nth-child(2){animation-delay:.18s}
        .boot-dot:nth-child(3){animation-delay:.36s}
        @keyframes bdot{0%,80%,100%{transform:scale(.66);opacity:.45}40%{transform:scale(1);opacity:1}}
        .boot-msg{min-height:20px;font-size:13.5px;font-weight:600;color:var(--ink2)}
        .boot-sub{margin-top:12px;font-size:11.5px;color:var(--ink3);line-height:1.55}
        .boot-card--err .boot-mark{background:linear-gradient(135deg,#7f1d1d,#ef4444);
            box-shadow:0 10px 24px -8px rgba(239,68,68,.55)}
        .boot-err-msg{font-size:13px;color:var(--ink2);line-height:1.6;margin-top:8px}
        .boot-note{font-style:italic;font-size:11.5px;color:var(--ink3);margin-top:14px}
        .boot-retry{display:inline-block;margin-top:18px;padding:10px 22px;
            border-radius:var(--r-sm);background:var(--jpm);color:#fff!important;
            font-size:13px;font-weight:700;text-decoration:none;
            box-shadow:0 2px 10px -2px rgba(0,94,184,.6);transition:all .15s ease}
        .boot-retry:hover{background:var(--jpm-dark);transform:translateY(-1px)}

        /* ---------- responsive ---------- */
        @media (max-width:1180px){
            .exec-grid{grid-template-columns:repeat(3,minmax(0,1fr))}
            .paired-row{flex-direction:column}
            .insight-col{flex:1 1 auto;min-width:0}
            .chart-col .card,.insight-col .card{min-height:0}
            .peer-metric-wrap{flex:1 1 100%;max-width:none;min-width:0}
        }
        @media (max-width:680px){
            .hdr-inner,.main-wrap,.ftr{padding-left:16px;padding-right:16px}
            .exec-grid{grid-template-columns:repeat(2,minmax(0,1fr))}
            .det-row{grid-template-columns:minmax(150px,1.6fr) minmax(80px,1fr) minmax(76px,.8fr);}
            .det-spark,.det-col-yoy{display:none}
            .boot-card{padding:30px 22px}
        }
        @media (prefers-reduced-motion:reduce){
            *,*::before,*::after{animation-duration:.001s!important;
                transition-duration:.001s!important}
        }
    </style>
</head>
<body>
    {%app_entry%}
    <footer>
        {%config%}
        {%scripts%}
        {%renderer%}
    </footer>
</body>
</html>"""


# =============================================================================
# DASHBOARD BUILDER -- layout + figure/insight factories. Callbacks live at
# module level (register_callbacks) so they can be registered before data
# exists; every method here is pure given self.df.
# =============================================================================
class DashboardBuilder:
    def __init__(self, df, cecl=None, missing_banks=None):
        self.df = df
        self.GHB = PRIMARY_BANK_DISPLAY_NAME
        self.cecl = cecl or {}
        self.missing_banks = missing_banks or []
        self.metrics = [m for m in METRIC_ORDER if m in df.columns]
        present = set(df['Bank'].unique())
        self.banks = [b['display'] for b in BANK_INFO if b['display'] in present]
        self.peers = [b for b in self.banks if b != self.GHB]
        prim = df[df['Bank'] == self.GHB]
        self.dates = sorted(prim['Date'].unique(), reverse=True) if not prim.empty \
            else sorted(df['Date'].unique(), reverse=True)
        self.latest = pd.Timestamp(self.dates[0]) if len(self.dates) else None
        self.default_metric = ('Return on Assets' if 'Return on Assets' in self.metrics
                               else (self.metrics[0] if self.metrics else None))
        # Per-bank render caches (cleared only on process restart -- data is static).
        self._bframes = {}
        self._bd_cache = {}

    # ------------------------------------------------------------ small utils
    def _metric_option(self, m):
        return {'label': m, 'value': m}

    def _mdd(self, did, value, options, multi=False, clearable=False, placeholder=None):
        return dcc.Dropdown(id=did, options=options, value=value, multi=multi,
                            clearable=clearable, placeholder=placeholder,
                            className='dd')

    def _date_options(self):
        return [{'label': pd.Timestamp(d).strftime('%m/%d/%Y'), 'value': str(pd.Timestamp(d))}
                for d in self.dates]

    def _fmt(self, v, m):
        return fmt_val(v, m, with_unit=True)

    def _bank_frame(self, bank):
        if bank not in self._bframes:
            self._bframes[bank] = self.df[self.df['Bank'] == bank].sort_values('Date')
        return self._bframes[bank]

    def _bank_qoq_yoy(self, bank, metric, date):
        return compute_period_deltas(self.df, bank, metric, date)

    def _bank_spark(self, bank, metric, lookback=8):
        return get_sparkline_series(self.df, bank, metric, lookback)

    def _window_bounds(self, banks, years):
        sub = self.df[self.df['Bank'].isin(banks)]
        if sub.empty:
            return None, None
        end = sub['Date'].max()
        start = end - pd.DateOffset(years=years)
        return start, end

    def _window_label(self, start, end):
        if start is None or end is None:
            return ""
        return f"{start.strftime('%m/%Y')}\u2013{end.strftime('%m/%Y')}"

    def _rdef(self, m, prefix=None):
        if not m:
            return html.Div()
        txt = METRIC_DEFINITIONS.get(m, "No definition available.")
        label = f"{prefix}: {m}" if prefix else m
        return html.Div([html.Span(label + " \u2014 ", style={'fontWeight': 700}),
                         html.Span(txt)], className='peer-def')

    def _peer_metric_definition(self, m):
        return self._rdef(m)

    def _ef(self, msg=""):
        fig = go.Figure()
        fig.update_layout(template='plotly_white', height=PAIRED_GRAPH_HEIGHT,
                          margin=dict(l=10, r=10, t=10, b=10),
                          xaxis={'visible': False}, yaxis={'visible': False},
                          annotations=[dict(text=msg or "", showarrow=False,
                                            font=dict(size=13, color=CS['text3']))])
        return self._bl(fig)

    def _bl(self, fig):
        """Lock the chart view: hover-only interaction, no zoom/pan artifacts."""
        fig.update_layout(dragmode=False, hoverlabel=dict(
            bgcolor='white', bordercolor=CS['border_strong'],
            font=dict(family='Inter, sans-serif', size=12, color=CS['text'])))
        fig.update_xaxes(fixedrange=True)
        fig.update_yaxes(fixedrange=True)
        return fig

    def _base_fig_layout(self, fig, height=PAIRED_GRAPH_HEIGHT):
        fig.update_layout(
            template='plotly_white', height=height,
            margin=dict(l=14, r=14, t=12, b=12),
            font=dict(family='Inter, sans-serif', size=11.5, color=CS['text2']),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            legend=dict(orientation='h', yanchor='bottom', y=1.01, x=0,
                        font=dict(size=10.5)),
        )
        fig.update_xaxes(gridcolor=CS['grid'], zeroline=False)
        fig.update_yaxes(gridcolor=CS['grid'], zeroline=False)
        return fig

    # -------------------------------------------------------- header sections
    def _exec_banner(self, selected_peers):
        cohort = [self.GHB] + [p for p in (selected_peers or []) if p in self.peers]
        cards = []
        for metric, cat in EXECUTIVE_KPIS:
            if metric not in self.metrics or self.latest is None:
                continue
            row = self.df[(self.df['Bank'] == self.GHB) & (self.df['Date'] == self.latest)]
            val = row.iloc[0][metric] if not row.empty else None
            qoq, _yoy = self._bank_qoq_yoy(self.GHB, metric, self.latest)
            d_txt, d_col = fmt_delta(val, qoq, metric)
            chip_cls = 'flat'
            if d_col == CS['good']:
                chip_cls = 'up'
            elif d_col == CS['bad']:
                chip_cls = 'down'
            rank, total, _pct = compute_peer_rank(self.df, self.GHB, metric, self.latest, cohort)
            rank_txt = f"#{rank} of {total}" if rank else "\u2014"
            spark = self._bank_spark(self.GHB, metric)
            cards.append(html.Div([
                html.Div(cat, className='exec-kpi-cat'),
                html.Div(metric, className='exec-kpi-label'),
                html.Div(self._fmt(val, metric), className='exec-kpi-val'),
                html.Div([
                    html.Span(d_txt + " QoQ", className=f'delta-chip {chip_cls}'),
                    html.Span(rank_txt, className='exec-rank'),
                ], className='exec-kpi-row'),
                make_sparkline_img_cached(spark),
            ], className='exec-card'))
        snap = self.latest.strftime('%m/%d/%Y') if self.latest is not None else "\u2014"
        return html.Div([
            html.Div([
                html.H3("Executive Snapshot \u2014 " + self.GHB, className='sec-title'),
                html.P(f"Latest quarter {snap} \u00b7 QoQ deltas \u00b7 rank vs "
                       f"{len(cohort) - 1} selected peers", className='sec-sub'),
            ], className='sec-head'),
            html.Div(cards, className='exec-grid'),
        ])

    def _missing_data_banner(self):
        notes = []
        if self.missing_banks:
            notes.append(f"FDIC returned no data for: {', '.join(self.missing_banks)}. "
                         f"They are excluded from peer statistics this session.")
        cov = self.cecl.get('coverage_pct') if self.cecl else None
        if cov is not None:
            notes.append(f"CECL concentration adjustment active: add-back applied to "
                         f"{self.cecl['applied_rows']} of {self.cecl['window_rows']} "
                         f"bank-quarters in the 2020\u20132024 window ({cov:.0f}% coverage).")
        if not notes:
            return html.Div()
        return html.Div(" ".join(notes), className='warn-banner')

    # ---------------------------------------------------------- reference card
    def _reference_section(self):
        sections = []
        for cat_name, cat_metrics in METRIC_CATEGORIES:
            present = [m for m in cat_metrics if m in self.metrics]
            if not present:
                continue
            rows = []
            for m in present:
                txt = METRIC_DEFINITIONS.get(m, "")
                rows.append(html.Div([
                    html.Div(m, className="ref-name"),
                    html.Div(txt, className="ref-desc"),
                ], className="ref-row"))
            accent = CATEGORY_ACCENTS.get(cat_name, CS['primary'])
            sections.append(html.Details([
                html.Summary([
                    html.Span(className='ref-accent', style={'background': accent}),
                    html.Span(CATEGORY_SHORT_LABELS.get(cat_name, cat_name),
                              className='ref-cat-label'),
                    html.Span(f"{len(present)}", className='ref-cat-count'),
                    html.Span("\u25B6", className='ref-chev'),
                ], className='ref-summary'),
                html.Div(rows, className="ref-body"),
            ], className="ref-cat"))
        return html.Div([
            html.Div([
                html.H6("Metric Reference Guide", className="ct"),
                html.Span(f"{len(self.metrics)} metrics across "
                          f"{len(METRIC_CATEGORIES)} categories \u00b7 click a category to expand",
                          className="rng"),
            ], className="ch"),
            html.Div(sections, className="ref-wrap"),
        ], className="card ref-card sec")

    # ----------------------------------------------------------------- layout
    def _layout(self):
        metric_opts = [self._metric_option(m) for m in self.metrics]
        date_opts = self._date_options()
        latest_val = date_opts[0]['value'] if date_opts else None
        peer_opts = [{'label': p, 'value': p} for p in self.peers]
        bank_opts = [{'label': b, 'value': b} for b in self.banks]
        year_opts = [{'label': f"{y} yr" if y > 1 else "1 yr", 'value': y}
                     for y in (1, 2, 3, 5, 10, 23)]

        header = html.Div(html.Div([
            html.H1(DASHBOARD_TITLE, className='hdr-title'),
            html.Div([
                html.Span([html.Span(className='live-dot'), "FDIC live"],
                          className='hdr-live'),
                html.Span(f"{len(self.banks)} institutions \u00b7 "
                          f"{len(self.metrics)} UBPR-aligned metrics \u00b7 "
                          f"quarterly since {REQUESTED_START_DATE_DISPLAY}"),
                html.Span(f"Latest data: "
                          f"{self.latest.strftime('%m/%d/%Y') if self.latest is not None else '\u2014'}"),
            ], className='hdr-meta'),
            html.Div(HEADER_DISCLOSURE_SHORT, className='hdr-disc'),
        ], className='hdr-inner'), className='hdr')

        peer_card = html.Div([
            html.Div([
                html.H6("Peer Set", className='ct'),
                html.Span(PEER_UNIVERSE_LABEL, className='csub'),
            ], className='ch'),
            html.Div([
                html.Div(self._mdd('peer-sel', self.peers, peer_opts, multi=True,
                                   placeholder="Select peer banks\u2026"),
                         className='peer-dd-wrap'),
                html.Div([
                    html.Button("Select all", id='sel-all', n_clicks=None, className='btn-mini'),
                    html.Button("Clear", id='sel-clear', n_clicks=None, className='btn-mini'),
                ], className='peer-actions'),
            ], className='peer-row'),
            html.Div(f"{self.GHB} is always included; statistics use the selected peers.",
                     className='peer-count'),
        ], className='card peer-card')

        snapshot_sec = html.Div([
            html.Div([
                html.H3("Peer Comparison", className='sec-title'),
                html.P("Point-in-time snapshot and trailing trend vs the selected peer set.",
                       className='sec-sub'),
            ], className='sec-head'),
            html.Div([
                html.Div([
                    html.Span("Metric", className='ctl-label'),
                    self._mdd('peer-metric', self.default_metric, metric_opts),
                ], id='peer-metric-wrap', className='ctl peer-metric-wrap'),
                html.Div([
                    html.Span("Snapshot date", className='ctl-label'),
                    self._mdd('r1d', latest_val, date_opts),
                ], className='ctl', style={'flex': '0 0 190px'}),
                html.Div([
                    html.Span("Trend window", className='ctl-label'),
                    self._mdd('r2t', 5, year_opts),
                ], className='ctl', style={'flex': '0 0 130px'}),
                html.Span(id='r2r', className='range-label',
                          style={'marginLeft': 'auto', 'alignSelf': 'flex-end'}),
            ], className='control-row'),
            html.Div(id='peer-def'),
            html.Div([
                html.Div(html.Div([
                    html.Div([html.H6("Peer Snapshot", className='ct'),
                              html.Span("ranked bar \u00b7 JPM highlighted", className='csub')],
                             className='ch'),
                    dcc.Loading(dcc.Graph(id='r1c', config=GRAPH_CONFIG), type='dot',
                                color=CS['primary']),
                ], className='card'), className='chart-col'),
                html.Div(html.Div([
                    html.Div([html.H6("Metric Overview", className='ct')], className='ch'),
                    html.Div(id='r1o', className='insight-shell'),
                ], className='card'), className='insight-col'),
            ], className='paired-row'),
            html.Div([
                html.Div(html.Div([
                    html.Div([html.H6("Peer Trend", className='ct'),
                              html.Span("peer band = min\u2013max \u00b7 dashed = peer median",
                                        className='csub')], className='ch'),
                    dcc.Loading(dcc.Graph(id='r2c', config=GRAPH_CONFIG), type='dot',
                                color=CS['primary']),
                ], className='card'), className='chart-col'),
                html.Div(html.Div([
                    html.Div([html.H6("Trend Analysis", className='ct')], className='ch'),
                    html.Div(id='r2a', className='insight-shell'),
                ], className='card'), className='insight-col'),
            ], className='paired-row', style={'marginTop': '16px'}),
        ], className='sec')

        dual_sec = html.Div([
            html.Div([
                html.H3("Metric Relationship", className='sec-title'),
                html.P("Two metrics for JPMorgan on independent axes, with Pearson correlation.",
                       className='sec-sub'),
            ], className='sec-head'),
            html.Div([
                html.Div([html.Span("Primary metric", className='ctl-label'),
                          self._mdd('r3p', self.default_metric, metric_opts)],
                         className='ctl', style={'flex': '1 1 320px', 'maxWidth': '480px'}),
                html.Div([html.Span("Secondary metric", className='ctl-label'),
                          self._mdd('r3s', 'Net Interest Margin'
                                    if 'Net Interest Margin' in self.metrics
                                    else self.default_metric, metric_opts)],
                         className='ctl', style={'flex': '1 1 320px', 'maxWidth': '480px'}),
                html.Div([html.Span("Window", className='ctl-label'),
                          self._mdd('r3t', 10, year_opts)],
                         className='ctl', style={'flex': '0 0 130px'}),
            ], className='control-row'),
            html.Div(id='r3f'),
            html.Div([
                html.Div(html.Div([
                    html.Div([html.H6("Dual-Axis Trend \u2014 JPMorgan", className='ct')],
                             className='ch'),
                    dcc.Loading(dcc.Graph(id='r3c', config=GRAPH_CONFIG), type='dot',
                                color=CS['primary']),
                ], className='card'), className='chart-col'),
                html.Div(html.Div([
                    html.Div([html.H6("JPMorgan Correlation Analysis", className='ct')],
                             className='ch'),
                    html.Div(id='r3x', className='insight-shell'),
                ], className='card jpm-corr-card'), className='insight-col'),
            ], className='paired-row'),
        ], className='sec')

        detail_sec = html.Div([
            html.Div([
                html.H6("All Metrics", id='det-title', className='ct'),
                html.Span("every metric \u00b7 8-quarter sparkline \u00b7 QoQ / YoY deltas",
                          className='csub'),
            ], className='ch'),
            html.Div([
                html.Div([html.Span("Bank", className='ctl-label'),
                          self._mdd('det-bank', self.GHB, bank_opts)],
                         id='det-bank-wrap', className='ctl det-bank-wrap'),
                html.Div([html.Span("As of", className='ctl-label'),
                          self._mdd('det-date', latest_val, date_opts)],
                         className='ctl det-date-wrap'),
                html.Button("Export all periods (Excel)", id='exp', n_clicks=None,
                            className='btn-export'),
                dcc.Download(id='dl'),
                html.Span("Deltas: QoQ vs prior quarter \u00b7 YoY vs same quarter last year",
                          className='det-legend'),
            ], className='det-controls'),
            dcc.Loading(html.Div(id='det'), type='dot', color=CS['primary']),
        ], className='card sec det-card')

        footer = html.Div([
            html.Div(FOOTER_DISCLOSURE_NOTE),
            html.Div(f"Source: FDIC BankFind Suite API \u00b7 {BASE_URL} \u00b7 "
                     f"UBPR-aligned methodology \u00b7 values as reported, never synthesized.",
                     style={'marginTop': '4px'}),
        ], className='ftr')

        return html.Div([
            header,
            html.Div([
                peer_card,
                self._missing_data_banner(),
                html.Div(self._exec_banner(self.peers), id='exec-banner-wrap',
                         className='sec'),
                snapshot_sec,
                dual_sec,
                detail_sec,
                self._reference_section(),
            ], className='main-wrap'),
            footer,
        ])

    # =========================================================== chart methods
    def _bar(self, f, m, dt=None):
        """Ranked horizontal peer snapshot. JPM gets the brand color and a dark
        outline; peers are muted slate. Direction-aware: best at the TOP."""
        f = f.dropna(subset=[m]).drop_duplicates(subset=['Bank'], keep='last')
        if f.empty:
            return self._ef("No data for this metric/date")
        asc = is_inverse_metric(m)
        f = f.sort_values(m, ascending=asc)
        banks = list(f['Bank'])
        vals = list(f[m])
        colors = [CS['ghb'] if b == self.GHB else CS['peer'] for b in banks]
        line_colors = [CS['primary_dark'] if b == self.GHB else 'rgba(0,0,0,0)' for b in banks]
        line_widths = [1.4 if b == self.GHB else 0 for b in banks]
        texts = [self._fmt(v, m) for v in vals]
        fig = go.Figure(go.Bar(
            x=vals, y=banks, orientation='h',
            marker=dict(color=colors, opacity=[1.0 if b == self.GHB else CS['peer_op']
                                               for b in banks],
                        line=dict(color=line_colors, width=line_widths)),
            text=texts, textposition='outside', cliponaxis=False,
            textfont=dict(size=10.5, family='Inter, sans-serif'),
            hovertemplate='%{y}: %{text}<extra></extra>',
        ))
        peer_vals = [v for b, v in zip(banks, vals) if b != self.GHB]
        if peer_vals:
            avg = float(np.nanmean(peer_vals))
            fig.add_vline(x=avg, line_dash='dot', line_color=CS['peer_band_mid'],
                          line_width=1.2,
                          annotation_text=f"peer avg {self._fmt(avg, m)}",
                          annotation_position='top',
                          annotation_font=dict(size=10, color=CS['text3']))
        self._base_fig_layout(fig, height=max(PAIRED_GRAPH_HEIGHT, 22 * len(banks) + 60))
        fig.update_layout(showlegend=False, yaxis=dict(autorange='reversed'),
                          margin=dict(l=14, r=64, t=26, b=12))
        if is_dollar_metric(m):
            fig.update_xaxes(tickformat='~s')
        return self._bl(fig)

    def _ov(self, f, m, dt):
        """Metric overview panel: JPM value, percentile gauge, rank, peer
        statistics, deltas, and the 8-quarter JPM momentum sparkline."""
        f = f.dropna(subset=[m]).drop_duplicates(subset=['Bank'], keep='last')
        row = f[f['Bank'] == self.GHB]
        val = row.iloc[0][m] if not row.empty else None
        cohort = list(f['Bank'])
        rank, total, pct = compute_peer_rank(self.df, self.GHB, m, dt, cohort)
        peer_df = f[f['Bank'] != self.GHB]
        stats_rows = []
        if not peer_df.empty:
            pv = peer_df[m].dropna()
            if not pv.empty:
                hi_i, lo_i = pv.idxmax(), pv.idxmin()
                stats_rows = [
                    ("Peer median", self._fmt(float(pv.median()), m)),
                    ("Peer average", self._fmt(float(pv.mean()), m)),
                    ("Peer high", f"{self._fmt(float(pv.loc[hi_i]), m)} \u00b7 "
                                  f"{peer_df.loc[hi_i, 'Bank']}"),
                    ("Peer low", f"{self._fmt(float(pv.loc[lo_i]), m)} \u00b7 "
                                 f"{peer_df.loc[lo_i, 'Bank']}"),
                ]
        qoq, yoy = self._bank_qoq_yoy(self.GHB, m, dt)
        q_txt, q_col = fmt_delta(val, qoq, m)
        y_txt, y_col = fmt_delta(val, yoy, m)

        def chip(txt, col, label):
            cls = 'up' if col == CS['good'] else ('down' if col == CS['bad'] else 'flat')
            return html.Span(f"{txt} {label}", className=f'delta-chip {cls}',
                             style={'marginRight': '6px'})

        rank_line = (f"Rank #{rank} of {total} \u00b7 "
                     f"{'higher' if not is_inverse_metric(m) else 'lower'} is better"
                     if rank else "Rank unavailable")
        return html.Div([
            html.Div([
                html.Div([
                    html.Div([html.Span(self._fmt(val, m), className='ov-val'),
                              html.Span(self.GHB, className='ov-unit')]),
                    html.Div(rank_line, className='ov-rank'),
                    html.Div([chip(q_txt, q_col, 'QoQ'), chip(y_txt, y_col, 'YoY')],
                             style={'marginTop': '9px'}),
                ]),
                make_percentile_arc_img(pct),
            ], className='ov-top'),
            html.Div([html.Div([html.Div(lbl, className='ov-stat-label'),
                                html.Div(v, className='ov-stat-val')],
                               className='ov-stat') for lbl, v in stats_rows],
                     className='ov-stats'),
            html.Div("JPM Momentum \u00b7 trailing 8 quarters", className='ov-mom-label'),
            make_sparkline_img_cached(self._bank_spark(self.GHB, m), width=240, height=44,
                                      cls='spark-img'),
        ], className='ov-wrap')

    def _trend(self, banks, m, years):
        start, end = self._window_bounds(banks, years)
        if start is None:
            return self._ef("No data")
        sub = self.df[(self.df['Bank'].isin(banks)) & (self.df['Date'] >= start)
                      & (self.df['Date'] <= end)]
        sub = sub.dropna(subset=[m])
        if sub.empty:
            return self._ef("No data in window")
        piv = (sub.drop_duplicates(subset=['Bank', 'Date'], keep='last')
                  .pivot(index='Date', columns='Bank', values=m).sort_index())
        peer_cols = [c for c in piv.columns if c != self.GHB]
        fig = go.Figure()
        if peer_cols:
            pmax = piv[peer_cols].max(axis=1)
            pmin = piv[peer_cols].min(axis=1)
            pmed = piv[peer_cols].median(axis=1)
            fig.add_trace(go.Scatter(x=piv.index, y=pmax, mode='lines',
                                     line=dict(width=0), hoverinfo='skip',
                                     showlegend=False))
            fig.add_trace(go.Scatter(x=piv.index, y=pmin, mode='lines',
                                     line=dict(width=0), fill='tonexty',
                                     fillcolor=CS['peer_tint'], hoverinfo='skip',
                                     name='Peer range', showlegend=True))
            fig.add_trace(go.Scatter(x=piv.index, y=pmed, mode='lines',
                                     name='Peer median',
                                     line=dict(color=CS['peer_band_mid'], width=1.4,
                                               dash='dash'),
                                     hovertemplate='Peer median: %{y:.2f}<extra></extra>'))
        if self.GHB in piv.columns:
            jp = piv[self.GHB]
            fig.add_trace(go.Scatter(x=piv.index, y=jp, mode='lines', name=self.GHB,
                                     line=dict(color=CS['ghb'], width=2.6),
                                     hovertemplate=self.GHB + ': %{y:.2f}<extra></extra>'))
            jnn = jp.dropna()
            if not jnn.empty:
                fig.add_trace(go.Scatter(x=[jnn.index[-1]], y=[jnn.iloc[-1]],
                                         mode='markers', showlegend=False,
                                         marker=dict(size=7, color=CS['ghb'],
                                                     line=dict(color='white', width=2)),
                                         hoverinfo='skip'))
        self._base_fig_layout(fig)
        fig.update_layout(hovermode='x unified')
        if is_dollar_metric(m):
            fig.update_yaxes(tickformat='~s')
        return self._bl(fig)

    def _ta(self, banks, m, years):
        start, end = self._window_bounds(banks, years)
        if start is None:
            return html.P("No data", className='emp')
        jdf = self._bank_frame(self.GHB)
        jdf = jdf[(jdf['Date'] >= start) & (jdf['Date'] <= end)].dropna(subset=[m])
        if len(jdf) < 2:
            return html.P("Not enough history in this window", className='emp')
        vals = jdf[m].astype(float).values
        sv, ev = vals[0], vals[-1]
        chg = calc_trend_change(sv, ev, m)
        x = np.arange(len(vals))
        slope = float(np.polyfit(x, vals, 1)[0]) if len(vals) >= 2 else np.nan
        vol = float(np.nanstd(vals))
        mean = float(np.nanmean(vals))
        cv = (vol / abs(mean) * 100) if mean not in (0, None) and abs(mean) > 1e-12 else None
        # correlation with the peer median over the same dates
        corr_txt = "N/A"
        peers = [b for b in banks if b != self.GHB]
        if peers:
            sub = self.df[(self.df['Bank'].isin(peers)) & (self.df['Date'] >= start)
                          & (self.df['Date'] <= end)].dropna(subset=[m])
            if not sub.empty:
                pmed = (sub.drop_duplicates(subset=['Bank', 'Date'], keep='last')
                           .pivot(index='Date', columns='Bank', values=m)
                           .median(axis=1))
                joined = pd.concat([jdf.set_index('Date')[m], pmed], axis=1,
                                   join='inner').dropna()
                if len(joined) >= 3:
                    a = joined.iloc[:, 0].values
                    b = joined.iloc[:, 1].values
                    if np.std(a) > 1e-12 and np.std(b) > 1e-12:
                        r, _ = stats.pearsonr(a, b)
                        corr_txt = f"{r:+.2f}"
        items = [
            ("Window", self._window_label(start, end)),
            ("Start \u2192 end", f"{self._fmt(sv, m)} \u2192 {self._fmt(ev, m)}"),
            ("Net change", fmt_trend_change(chg, m)),
            ("Direction", trend_direction_label(slope)),
            ("Volatility (\u03c3)", f"{vol:,.2f}"),
            ("Coef. of variation", f"{cv:.1f}%" if cv is not None else "N/A"),
            ("Corr. w/ peer median", corr_txt),
            ("Observations", f"{len(vals)} quarters"),
        ]
        return html.Div([html.Div([html.Div(lbl, className='ov-stat-label'),
                                   html.Div(v, className='ov-stat-val')],
                                  className='ov-stat') for lbl, v in items],
                        className='ov-stats')

    def _dual(self, a, b, years):
        start, end = self._window_bounds([self.GHB], years)
        if start is None:
            return self._ef("No data")
        jdf = self._bank_frame(self.GHB)
        jdf = jdf[(jdf['Date'] >= start) & (jdf['Date'] <= end)]
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Scatter(x=jdf['Date'], y=jdf[a], name=a, mode='lines',
                                 line=dict(color=CS['ghb'], width=2.4),
                                 hovertemplate=a + ': %{y:.2f}<extra></extra>'),
                      secondary_y=False)
        fig.add_trace(go.Scatter(x=jdf['Date'], y=jdf[b], name=b, mode='lines',
                                 line=dict(color=CS['warn'], width=2.2, dash='dot'),
                                 hovertemplate=b + ': %{y:.2f}<extra></extra>'),
                      secondary_y=True)
        self._base_fig_layout(fig)
        fig.update_layout(hovermode='x unified')
        fig.update_yaxes(title_text=a, title_font=dict(size=10.5), secondary_y=False)
        fig.update_yaxes(title_text=b, title_font=dict(size=10.5), secondary_y=True,
                         gridcolor='rgba(0,0,0,0)')
        if is_dollar_metric(a):
            fig.update_yaxes(tickformat='~s', secondary_y=False)
        if is_dollar_metric(b):
            fig.update_yaxes(tickformat='~s', secondary_y=True)
        return self._bl(fig)

    def _corr(self, a, b, years):
        start, end = self._window_bounds([self.GHB], years)
        if start is None:
            return html.P("No data", className='emp')
        jdf = self._bank_frame(self.GHB)
        jdf = jdf[(jdf['Date'] >= start) & (jdf['Date'] <= end)][[a, b]].dropna()
        if len(jdf) < 3:
            return html.P("Fewer than 3 overlapping quarters \u2014 correlation not computed.",
                          className='emp')
        x, y = jdf[a].values.astype(float), jdf[b].values.astype(float)
        if np.std(x) < 1e-12 or np.std(y) < 1e-12:
            return html.P("One series is constant in this window \u2014 correlation undefined.",
                          className='emp')
        r, p = stats.pearsonr(x, y)
        ar = abs(r)
        strength = ("very strong" if ar >= 0.8 else "strong" if ar >= 0.6 else
                    "moderate" if ar >= 0.4 else "weak" if ar >= 0.2 else "negligible")
        direction = "positive" if r > 0 else "negative"
        col = CS['good'] if r > 0.2 else (CS['bad'] if r < -0.2 else CS['neutral'])
        interp = (f"A {strength} {direction} relationship between \u201c{a}\u201d and "
                  f"\u201c{b}\u201d for {self.GHB} across {len(jdf)} quarters "
                  f"({self._window_label(start, end)}). p = {p:.3f}. "
                  f"Correlation is descriptive, not causal.")
        return html.Div([
            html.Div("Pearson r", className='corr-label'),
            html.Div(f"{r:+.2f}", className='corr-val', style={'color': col}),
            html.Div(interp, className='corr-interp'),
        ])

    # ----------------------------------------------------- All-Metrics detail
    def _bd(self, dt, bank=None):
        bank = bank or self.GHB
        key = (bank, pd.Timestamp(dt))
        if key in self._bd_cache:
            return self._bd_cache[key]
        out = self._bd_build(pd.Timestamp(dt), bank)
        self._bd_cache[key] = out
        return out

    def _bd_build(self, dt, bank):
        bf = self._bank_frame(bank)
        row = bf[bf['Date'] == dt]
        if row.empty:
            return html.P(f"No data for {bank} on {dt.strftime('%m/%d/%Y')}.",
                          className='emp')
        r = row.iloc[0]
        cats = []
        for cat_name, cat_metrics in METRIC_CATEGORIES:
            present = [m for m in cat_metrics if m in self.metrics]
            if not present:
                continue
            accent = CATEGORY_ACCENTS.get(cat_name, CS['primary'])
            bg = CATEGORY_BG.get(cat_name, '#f8fafc')
            rows = [html.Div([
                html.Div("Metric"), html.Div("Value", style={'textAlign': 'right'}),
                html.Div("8Q trend", className='det-spark'),
                html.Div("QoQ", style={'textAlign': 'right'}),
                html.Div("YoY", style={'textAlign': 'right'}, className='det-col-yoy'),
            ], className='det-row det-hdr-row')]
            for m in present:
                v = r[m]
                qoq, yoy = self._bank_qoq_yoy(bank, m, dt)
                q_txt, q_col = fmt_delta(v, qoq, m)
                y_txt, y_col = fmt_delta(v, yoy, m)
                rows.append(html.Div([
                    html.Div(m, className='det-name', title=METRIC_DEFINITIONS.get(m, '')),
                    html.Div(self._fmt(v, m), className='det-val'),
                    html.Div(make_sparkline_img_cached(
                        self._bank_spark(bank, m)), className='det-spark'),
                    html.Div(q_txt, className='det-delta', style={'color': q_col}),
                    html.Div(y_txt, className='det-delta det-col-yoy',
                             style={'color': y_col}),
                ], className='det-row'))
            cats.append(html.Div([
                html.Div([
                    html.Span(style={'width': '4px', 'height': '16px',
                                     'borderRadius': '3px', 'background': accent,
                                     'display': 'inline-block'}),
                    html.Span(CATEGORY_SHORT_LABELS.get(cat_name, cat_name)),
                    html.Span(f"{len(present)} metrics", className='det-cat-count'),
                ], className='det-cat-head', style={'background': bg, 'color': accent}),
                html.Div(rows),
            ], className='det-cat'))
        return html.Div(cats)


# =============================================================================
# NON-BLOCKING BOOT -- module-level app binds to $PORT instantly; the FDIC
# load runs in a background daemon thread started lazily on the FIRST page
# request (lazy start also keeps the debug reloader and gunicorn forking from
# double-spawning it). Visitors see a live-updating boot screen that reloads
# itself into the dashboard when the data is ready.
# =============================================================================
class _AppState:
    def __init__(self):
        self.lock = threading.Lock()
        self.thread_started = False
        self.status = 'loading'      # loading | ready | error
        self.message = 'Contacting the FDIC BankFind API\u2026'
        self.error_title = None
        self.error_message = None
        self.builder = None
        self.missing_banks = []


STATE = _AppState()


def _load_data():
    def msg(m):
        STATE.message = m
        logger.info(f"[boot] {m}")

    try:
        sd = DEFAULT_START_DATE
        ed = datetime.today().strftime('%Y%m%d')  # recompute: dyno may outlive import day
        msg(f"Requesting quarterly filings for {len(BANK_INFO)} institutions\u2026")
        service = BankDataService()
        df = service.get_metrics_data(sd, ed, progress=msg)
        if df is None or df.empty:
            raise FDICDataUnavailableError(
                "The FDIC API responded but produced zero usable bank-quarters.")
        msg("Assembling peer statistics and layout\u2026")
        present = set(df['Bank'].unique())
        missing = [b['display'] for b in BANK_INFO if b['display'] not in present]
        if PRIMARY_BANK_DISPLAY_NAME not in present:
            raise FDICDataUnavailableError(
                f"{PRIMARY_BANK_DISPLAY_NAME} (cert {PRIMARY_BANK_CERT}) is missing "
                f"from the dataset; the dashboard cannot render without its anchor bank.")
        cecl = service.calc.cecl_status()
        builder = DashboardBuilder(df, cecl=cecl, missing_banks=missing)
        STATE.missing_banks = missing
        STATE.builder = builder          # set builder BEFORE flipping status
        STATE.status = 'ready'
        logger.info(f"[boot] READY: {len(builder.banks)} banks, "
                    f"{len(builder.metrics)} metrics, latest {builder.latest}. "
                    f"Missing: {missing or 'none'}. CECL: {cecl}")
    except FDICDataUnavailableError as e:
        STATE.error_title = "FDIC Data Unavailable"
        STATE.error_message = str(e)
        STATE.status = 'error'
        logger.error(f"[boot] FDIC data unavailable: {e}")
    except Exception as e:  # noqa: BLE001 -- surface anything to the error page
        STATE.error_title = "Dashboard Error"
        STATE.error_message = f"{type(e).__name__}: {e}"
        STATE.status = 'error'
        logger.exception("[boot] Unhandled error while building the dashboard.")


def _ensure_loader_started():
    if STATE.thread_started:
        return
    with STATE.lock:
        if STATE.thread_started:
            return
        STATE.thread_started = True
        t = threading.Thread(target=_load_data, name='fdic-loader', daemon=True)
        t.start()
        logger.info("[boot] FDIC loader thread started.")


def _loading_layout():
    return html.Div([
        dcc.Location(id='boot-loc', refresh=True),
        dcc.Interval(id='boot-int', interval=1250),
        html.Div([
            html.Div("JPM", className='boot-mark'),
            html.Div(DASHBOARD_TITLE, className='boot-title'),
            html.Div([html.Span(className='boot-dot'), html.Span(className='boot-dot'),
                      html.Span(className='boot-dot')], className='boot-dots'),
            html.Div(STATE.message, id='boot-msg', className='boot-msg'),
            html.Div(f"Pulling full quarterly history for {len(BANK_INFO)} systemically "
                     f"important banks from the FDIC BankFind API. A cold start takes "
                     f"~30\u201390 seconds \u00b7 this page refreshes automatically.",
                     className='boot-sub'),
        ], className='boot-card'),
    ], className='boot-screen')


def _error_layout(title, message):
    return html.Div([
        html.Div([
            html.Div("!", className='boot-mark'),
            html.Div(title or "Dashboard Error", className='boot-title'),
            html.Div(message or "An unexpected error occurred.", className='boot-err-msg'),
            html.Div("This dashboard renders real FDIC data only \u2014 no synthetic "
                     "fallback is ever substituted.", className='boot-note'),
            html.A("Retry data load", href='/?retry=1', className='boot-retry'),
        ], className='boot-card boot-card--err'),
    ], className='boot-screen')


def _serve_layout():
    # An explicit retry resets an errored loader and spawns a fresh attempt.
    try:
        if (STATE.status == 'error'
                and flask_request and flask_request.args.get('retry')):
            with STATE.lock:
                if STATE.status == 'error':
                    STATE.status = 'loading'
                    STATE.message = 'Retrying FDIC data load\u2026'
                    STATE.error_title = None
                    STATE.error_message = None
                    STATE.thread_started = False
    except RuntimeError:
        pass  # outside a request context (e.g., layout validation)
    _ensure_loader_started()
    if STATE.status == 'ready' and STATE.builder is not None:
        return STATE.builder._layout()
    if STATE.status == 'error':
        return _error_layout(STATE.error_title, STATE.error_message)
    return _loading_layout()


# =============================================================================
# APP + CALLBACKS (registered once at import; every callback guards on STATE)
# =============================================================================
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP],
                meta_tags=[{'name': 'viewport',
                            'content': 'width=device-width, initial-scale=1'}])
app.title = DASHBOARD_SHORT_TITLE
app.config.suppress_callback_exceptions = True
app.index_string = INDEX_STRING
app.layout = _serve_layout
server = app.server


def register_callbacks(app):

    @app.callback([Output('boot-msg', 'children'), Output('boot-loc', 'href')],
                  Input('boot-int', 'n_intervals'))
    def boot_poll(_n):
        if STATE.status in ('ready', 'error'):
            return no_update, '/'
        return STATE.message, no_update

    @app.callback(Output('peer-sel', 'value'),
                  [Input('sel-all', 'n_clicks'), Input('sel-clear', 'n_clicks')],
                  State('peer-sel', 'options'), prevent_initial_call=True)
    def sel_action(n_all, n_clear, options):
        ctx = dash.callback_context
        if not ctx.triggered:
            raise PreventUpdate
        trig = ctx.triggered[0]['prop_id'].split('.')[0]
        if trig == 'sel-all' and n_all:
            return [x['value'] for x in (options or [])]
        if trig == 'sel-clear' and n_clear:
            return []
        raise PreventUpdate

    @app.callback(Output('exec-banner-wrap', 'children'), Input('peer-sel', 'value'))
    def ue(p):
        b = STATE.builder
        if b is None:
            raise PreventUpdate
        return b._exec_banner(p or [])

    @app.callback(Output('peer-def', 'children'), Input('peer-metric', 'value'))
    def d_peer(m):
        b = STATE.builder
        if b is None:
            raise PreventUpdate
        return b._peer_metric_definition(m)

    @app.callback(Output('peer-metric-wrap', 'style'), Input('peer-metric', 'value'))
    def resize_peer_metric_control(m):
        base = {'flex': '0 0 500px', 'maxWidth': '610px', 'minWidth': '360px'}
        if STATE.builder is None or not m:
            return base
        # Widen the control for the long segment-metric names so the selected
        # value never truncates; capped to keep the control row balanced.
        width = min(610, max(360, 240 + 7 * len(str(m))))
        return {'flex': f'0 0 {width}px', 'maxWidth': '610px', 'minWidth': '360px'}

    @app.callback(Output('r3f', 'children'),
                  [Input('r3p', 'value'), Input('r3s', 'value')])
    def d3(a, b_):
        b = STATE.builder
        if b is None:
            raise PreventUpdate
        return html.Div([b._rdef(a, "Primary"), b._rdef(b_, "Secondary")])

    @app.callback([Output('r1c', 'figure'), Output('r1o', 'children')],
                  [Input('peer-metric', 'value'), Input('r1d', 'value'),
                   Input('peer-sel', 'value')])
    def u1(m, ds, p):
        b = STATE.builder
        if b is None:
            raise PreventUpdate
        if not m or not ds:
            return b._ef(""), html.Div()
        dt = pd.to_datetime(ds)
        bk = [b.GHB] + (p or [])
        f = b.df[(b.df['Date'] == dt) & b.df['Bank'].isin(bk)]
        if f.empty:
            return b._ef("No data"), html.Div()
        return b._bar(f, m, dt), b._ov(f, m, dt)

    @app.callback([Output('r2c', 'figure'), Output('r2a', 'children'),
                   Output('r2r', 'children')],
                  [Input('peer-metric', 'value'), Input('peer-sel', 'value'),
                   Input('r2t', 'value')])
    def u2(m, p, y):
        b = STATE.builder
        if b is None:
            raise PreventUpdate
        if not m:
            return b._ef(""), html.Div(), ""
        bk = [b.GHB] + (p or [])
        y = y or 5
        start, end = b._window_bounds(bk, y)
        return b._trend(bk, m, y), b._ta(bk, m, y), b._window_label(start, end)

    @app.callback([Output('r3c', 'figure'), Output('r3x', 'children')],
                  [Input('r3p', 'value'), Input('r3s', 'value'), Input('r3t', 'value')])
    def u3(a, b_, y):
        b = STATE.builder
        if b is None:
            raise PreventUpdate
        if not a or not b_:
            return b._ef(""), html.Div()
        return b._dual(a, b_, y or 10), b._corr(a, b_, y or 10)

    @app.callback(Output('det', 'children'),
                  [Input('det-date', 'value'), Input('det-bank', 'value')])
    def ud(ds, bank):
        b = STATE.builder
        if b is None:
            raise PreventUpdate
        if not ds:
            return html.P("Select a date", className='emp')
        return b._bd(pd.to_datetime(ds), bank or b.GHB)

    @app.callback(Output('dl', 'data'), Input('exp', 'n_clicks'),
                  State('det-bank', 'value'), prevent_initial_call=True)
    def export_all_periods(n, bank):
        b = STATE.builder
        if not n or b is None:
            raise PreventUpdate
        bank = bank or b.GHB
        payload = build_bank_export(b.df, bank_display=bank)
        safe = ''.join(ch if ch.isalnum() else '_' for ch in bank).strip('_')
        fname = f"{safe}_all_metrics_{datetime.today().strftime('%Y%m%d')}.xlsx"
        return dcc.send_bytes(payload, fname)


register_callbacks(app)


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 8050)))
